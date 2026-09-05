import csv
import hashlib
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from tools.analyze_drilling_workbooks import (
    analyze_vba_modules,
    assert_public_records_safe,
    build_calculation_rows,
    build_unit_rows,
    capture_catalog,
    classify_calculation_topics,
    compact_formula_families,
    convert_calamine_audit,
    detect_units,
    extract_formula_functions,
    extract_ooxml_workbook,
    extract_workbook_static,
    formula_has_external_reference,
    merge_captures,
    normalize_formula,
    parse_olevba_payload,
    public_formula_functions,
    validate_calamine_payload,
)


class FormulaAnalysisTests(unittest.TestCase):
    def test_formula_normalization_groups_copied_calculation_shapes(self):
        first = "=IF($A2>10,[Book.xlsx]Data!B2*C$1,0)"
        second = "=IF($A9>10,[Other.xlsm]Inputs!B9*C$1,0)"

        self.assertEqual(normalize_formula(first), normalize_formula(second))
        self.assertNotEqual(
            normalize_formula(first),
            normalize_formula("=IF($A9>25,[Other.xlsm]Inputs!B9*C$1,0)"),
        )
        self.assertEqual(extract_formula_functions(first), ["IF"])

    def test_formula_normalization_preserves_structured_columns_and_function_names(
        self,
    ):
        pressure = "=LOG10(Table1[Pressure])+[Book.xlsx]Data!A1"
        flow = "=LOG10(Table1[Flow])+[Book.xlsx]Data!A9"

        self.assertNotEqual(normalize_formula(pressure), normalize_formula(flow))
        self.assertIn("LOG10(", normalize_formula(pressure))
        self.assertTrue(formula_has_external_reference(pressure))
        self.assertFalse(formula_has_external_reference("=SUM(Table1[Pressure])"))
        self.assertEqual(
            public_formula_functions(["SUM", "Hydrodinamic"]), ["SUM", "UDF"]
        )

    def test_unit_detection_handles_compound_units_without_treating_in_as_word(self):
        text = "Hole diameter (in), mud weight 12.5 ppg, flow 500 gpm, temperature 180 deg F"

        units = {
            (row["canonical_unit"], row["dimension"]) for row in detect_units(text)
        }

        self.assertIn(("in", "length"), units)
        self.assertIn(("ppg", "density"), units)
        self.assertIn(("gal/min", "volumetric_flow"), units)
        self.assertIn(("degF", "temperature"), units)
        self.assertEqual(detect_units("Circulation in the annulus"), [])
        self.assertEqual(detect_units("N"), [])
        self.assertEqual(detect_units(r'0\ "K"', source_kind="number-format"), [])
        self.assertIn(
            ("N", "force"),
            {
                (row["canonical_unit"], row["dimension"])
                for row in detect_units("Axial force (N)")
            },
        )

    def test_unit_detection_preserves_prefixes_and_pressure_basis(self):
        units = {
            row["canonical_unit"]: row
            for row in detect_units(
                "Viscosity 2 mPa*s; torque 3 kN*m; heat capacity 4 kJ/(kg*K); "
                "pressures 5 psia and 6 psig"
            )
        }

        self.assertEqual(units["mPa*s"]["si_multiplier"], 0.001)
        self.assertEqual(units["kN*m"]["si_multiplier"], 1000.0)
        self.assertEqual(units["kJ/(kg*K)"]["si_multiplier"], 1000.0)
        self.assertEqual(units["psia"]["pressure_basis"], "absolute")
        self.assertEqual(units["psig"]["pressure_basis"], "gauge")

    def test_topics_are_multi_domain_and_include_reverse_thermal_flow(self):
        text = (
            "Reverse circulation heat exchanger outlet temperature, annular ECD, "
            "Reynolds number, nozzle pressure loss, and dogleg severity"
        )

        topics = set(classify_calculation_topics(text))

        self.assertTrue(
            {
                "thermal.reverse-circulation",
                "thermal.heat-exchange",
                "hydraulics.equivalent-circulating-density",
                "hydraulics.reynolds-number",
                "hydraulics.bit-nozzle",
                "directional.dogleg-severity",
            }.issubset(topics)
        )


class WorkbookExtractionTests(unittest.TestCase):
    def test_ooxml_parser_failure_uses_static_reader_fallback(self):
        fallback = {
            "status": "ok",
            "method": "calamine-static",
            "sheets": [],
            "formulas": [],
            "defined_names": [],
            "unit_mentions": [],
        }
        with (
            patch(
                "tools.analyze_drilling_workbooks._run_openxml_reader",
                side_effect=TypeError("unsupported workbook metadata"),
            ),
            patch(
                "tools.analyze_drilling_workbooks._run_calamine_reader",
                return_value=fallback,
            ) as static_reader,
        ):
            result = extract_workbook_static(
                Path("fixture.xlsm"),
                Path("reader"),
                None,
                Path("staging"),
                30,
            )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["method"], "calamine-static-ooxml-fallback")
        self.assertIn("TypeError", result["fallback_reason"])
        static_reader.assert_called_once()

    def test_encrypted_ooxml_uses_standard_office_decryption(self):
        fallback = {
            "status": "ok",
            "method": "calamine-static-after-standard-office-decryption",
            "sheets": [],
            "formulas": [],
            "defined_names": [],
            "unit_mentions": [],
        }
        with (
            patch(
                "tools.analyze_drilling_workbooks._run_openxml_reader",
                side_effect=RuntimeError("encrypted OOXML"),
            ),
            patch(
                "tools.analyze_drilling_workbooks._run_calamine_reader",
                side_effect=RuntimeError("Workbook is password protected"),
            ),
            patch(
                "tools.analyze_drilling_workbooks._run_calamine_after_standard_decryption",
                return_value=fallback,
            ) as decrypt,
        ):
            result = extract_workbook_static(
                Path("fixture.xlsm"),
                Path("reader"),
                Path("decryptor"),
                Path("staging"),
                30,
            )

        self.assertEqual(result["status"], "ok")
        decrypt.assert_called_once()

    def test_encrypted_binary_uses_standard_office_decryption(self):
        fallback = {
            "status": "ok",
            "method": "calamine-static-after-standard-office-decryption",
            "sheets": [],
            "formulas": [],
            "defined_names": [],
            "unit_mentions": [],
        }
        with (
            patch(
                "tools.analyze_drilling_workbooks._run_calamine_reader",
                side_effect=RuntimeError("Workbook is password protected"),
            ),
            patch(
                "tools.analyze_drilling_workbooks._run_calamine_after_standard_decryption",
                return_value=fallback,
            ) as decrypt,
        ):
            result = extract_workbook_static(
                Path("fixture.xls"),
                Path("reader"),
                Path("decryptor"),
                Path("staging"),
                30,
            )

        self.assertEqual(result["status"], "ok")
        self.assertIn("password protected", result["fallback_reason"])
        decrypt.assert_called_once()

    def test_ooxml_extraction_includes_hidden_sheets_formulas_units_and_names(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "fixture.xlsx"
            workbook = Workbook()
            hydraulics = workbook.active
            hydraulics.title = "Hydraulics"
            hydraulics["A1"] = "Flow rate (gpm)"
            hydraulics["B1"] = 500
            hydraulics["A2"] = "Pressure drop (psi)"
            hydraulics["B2"] = "=B1*2"
            hydraulics["B2"].number_format = '0.0 "psi"'
            hidden = workbook.create_sheet("Lookup")
            hidden.sheet_state = "veryHidden"
            hidden["A1"] = "=SUM(1,2)"
            workbook.defined_names.add(
                DefinedName("FlowRate", attr_text="'Hydraulics'!$B$1")
            )
            workbook.save(path)

            result = extract_ooxml_workbook(path)

        sheets = {row["sheet_name"]: row for row in result["sheets"]}
        formulas = {(row["sheet_name"], row["cell"]) for row in result["formulas"]}
        units = {
            row["canonical_unit"]
            for row in result["unit_mentions"]
            if row["sheet_name"] == "Hydraulics"
        }

        self.assertEqual(sheets["Lookup"]["visibility"], "veryHidden")
        self.assertEqual(sheets["Hydraulics"]["formula_cells"], 1)
        self.assertEqual(sheets["Lookup"]["formula_cells"], 1)
        self.assertEqual(formulas, {("Hydraulics", "B2"), ("Lookup", "A1")})
        self.assertTrue({"gal/min", "psi"}.issubset(units))
        self.assertEqual(result["defined_names"][0]["name"], "FlowRate")


class VbaAnalysisTests(unittest.TestCase):
    def test_vba_analysis_inventory_flags_events_calculations_and_process_launches(
        self,
    ):
        modules = [
            {
                "module_name": "Sheet1.cls",
                "code": (
                    "Private Sub Worksheet_Change(ByVal Target As Range)\n"
                    '  Shell "cmd.exe /c calc"\n'
                    "End Sub\n"
                    "Public Function PressureDrop(ByVal FlowRate As Double) As Double\n"
                    "  PressureDrop = FlowRate * 2\n"
                    "End Function\n"
                ),
            }
        ]

        procedures = analyze_vba_modules(modules)
        by_name = {row["procedure_name"]: row for row in procedures}

        self.assertEqual(
            by_name["Worksheet_Change"]["execution_trigger"], "worksheet-event"
        )
        self.assertIn("external-process", by_name["Worksheet_Change"]["risk_signals"])
        self.assertIn(
            "hydraulics.pressure-loss",
            by_name["PressureDrop"]["calculation_topics"],
        )
        self.assertEqual(by_name["PressureDrop"]["procedure_kind"], "function")

    def test_olevba_payload_is_reduced_to_procedures_and_indicators(self):
        payload = [
            {"type": "MetaInformation", "version": "0.60.2"},
            {
                "type": "OpenXML",
                "analysis": [
                    {
                        "type": "Suspicious",
                        "keyword": "Shell",
                        "description": "May run an executable file",
                    }
                ],
                "macros": [
                    {
                        "vba_filename": "Module1.bas",
                        "ole_stream": "VBA/Module1",
                        "code": (
                            "Public Function AnnularPressureLoss(FlowGpm As Double) As Double\n"
                            "  AnnularPressureLoss = FlowGpm * 2\n"
                            "End Function"
                        ),
                    }
                ],
            },
        ]

        result = parse_olevba_payload(payload)

        self.assertTrue(result["has_vba"])
        self.assertEqual(result["module_count"], 1)
        self.assertEqual(result["procedure_count"], 1)
        self.assertEqual(result["indicators"][0]["keyword"], "Shell")
        self.assertIn(
            "hydraulics.pressure-loss",
            result["procedures"][0]["calculation_topics"],
        )


class AggregationTests(unittest.TestCase):
    def test_calamine_contract_rejects_missing_schema_fields(self):
        with self.assertRaises(ValueError):
            validate_calamine_payload({"schema_version": "1.0.0", "sheets": []})

    def test_calamine_conversion_preserves_formula_context_and_text_occurrences(self):
        payload = {
            "schema_version": "1.0.0",
            "extension": "xls",
            "defined_names": [{"name": "FlowRate", "formula": "=Hydraulics!$B$1"}],
            "sheets": [
                {
                    "sheet_index": 1,
                    "sheet_name": "Hydraulics",
                    "sheet_kind": "worksheet",
                    "visibility": "hidden",
                    "rows_used": 3,
                    "columns_used": 2,
                    "populated_cells": 5,
                    "text_cells": [
                        {
                            "cells": ["A1"],
                            "text": "Flow rate (gpm)",
                            "occurrences": 3,
                        },
                        {
                            "cells": ["A2"],
                            "text": "Annular pressure loss (psi)",
                            "occurrences": 1,
                        },
                    ],
                    "formulas": [{"cell": "B2", "formula": "=B1*2"}],
                    "warnings": [],
                }
            ],
        }

        result = convert_calamine_audit(payload, "Legacy Hydraulics")

        self.assertEqual(result["sheets"][0]["formula_cells"], 1)
        self.assertEqual(result["formulas"][0]["cell"], "B2")
        self.assertIn(
            "hydraulics.pressure-loss",
            result["formulas"][0]["calculation_topics"],
        )
        unit_counts = {
            row["canonical_unit"]: row["occurrences"] for row in result["unit_mentions"]
        }
        self.assertEqual(unit_counts["gal/min"], 3)
        self.assertEqual(unit_counts["psi"], 1)

    def test_unit_rows_publish_si_semantics(self):
        rows = build_unit_rows(
            {"id": "a" * 16, "category": "hydraulics"},
            {
                "sheets": [{"sheet_name": "Inputs", "sheet_index": 1}],
                "unit_mentions": [
                    {
                        "sheet_name": "Inputs",
                        **detect_units("Pressure (psig)")[0],
                    }
                ],
            },
        )

        self.assertEqual(rows[0]["observed_unit"], "psig")
        self.assertEqual(rows[0]["canonical_si_unit"], "Pa")
        self.assertEqual(rows[0]["pressure_basis"], "gauge")
        self.assertEqual(rows[0]["sheet_id"], "s0001")

    def test_public_calculation_rows_group_formulas_without_exposing_formula_text(self):
        analysis = {
            "sheets": [
                {
                    "sheet_index": 1,
                    "sheet_name": "Hydraulics",
                    "sheet_kind": "worksheet",
                }
            ],
            "formulas": [
                {
                    "sheet_name": "Hydraulics",
                    "cell": "B2",
                    "formula": "=A2*2",
                    "formula_kind": "regular",
                    "family_id": "family-a",
                    "functions": ["Hydrodinamic"],
                    "volatile_functions": [],
                    "external_reference": False,
                    "calculation_topics": ["hydraulics.pressure-loss"],
                    "units": ["psi"],
                },
                {
                    "sheet_name": "Hydraulics",
                    "cell": "B3",
                    "formula": "=A3*3",
                    "formula_kind": "regular",
                    "family_id": "family-a",
                    "functions": ["Hydrodinamic"],
                    "volatile_functions": [],
                    "external_reference": False,
                    "calculation_topics": ["hydraulics.pressure-loss"],
                    "units": ["psi"],
                },
            ],
        }
        metadata = {"id": "abc123", "category": "hydraulics", "extension": ".xlsx"}

        rows = build_calculation_rows(metadata, analysis)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["occurrence_count"], 2)
        self.assertEqual(rows[0]["formula_family_id"], "f000001")
        self.assertEqual(rows[0]["functions"], "UDF")
        self.assertNotIn("formula", rows[0])
        self.assertNotIn("sheet_name", rows[0])
        assert_public_records_safe(rows)

    def test_repeated_formula_cells_compact_to_a_filled_column_family(self):
        analysis = {
            "sheets": [{"sheet_name": "Hydraulics", "sheet_index": 1}],
            "formulas": [
                {
                    "sheet_name": "Hydraulics",
                    "cell": cell,
                    "formula": f"=A{row}*2",
                    "normalized_formula": "=<REF>*<N>",
                    "formula_kind": "regular",
                    "family_id": "family-a",
                    "functions": [],
                    "volatile_functions": [],
                    "external_reference": False,
                    "calculation_topics": ["hydraulics.pressure-loss"],
                    "units": ["psi"],
                    "context_labels": ["Pressure loss (psi)"],
                }
                for row, cell in ((2, "B2"), (3, "B3"), (4, "B4"))
            ],
        }

        compacted = compact_formula_families(analysis)

        self.assertNotIn("formulas", compacted)
        family = compacted["formula_families"][0]
        self.assertEqual(family["occurrence_count"], 3)
        self.assertEqual(family["cell_ranges"], ["B2:B4"])
        self.assertEqual(family["layout_kind"], "filled-column")
        self.assertEqual(family["representative_formula"], "=A2*2")

    def test_public_record_gate_rejects_paths_and_raw_formulas(self):
        with self.assertRaises(ValueError):
            assert_public_records_safe([{"source_path": r"G:\My Drive\private.xls"}])
        with self.assertRaises(ValueError):
            assert_public_records_safe([{"raw_formula": "=A1+B1"}])
        for unsafe in (
            "+SUM(A1:A2)",
            "-2+3",
            "@SUM(A1:A2)",
            "user@example.com",
            "/home/user/private/book.xls",
            "file:///private/book.xls",
        ):
            with self.subTest(unsafe=unsafe), self.assertRaises(ValueError):
                assert_public_records_safe([{"value": unsafe}])

    def test_empty_capture_directory_cannot_replace_public_inventories(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "private" / "workbooks").mkdir(parents=True)
            with self.assertRaises(ValueError):
                merge_captures(root / "private", root / "public")

    def test_catalog_rejects_invalid_ids_and_hash_mismatches(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalog = root / "catalog"
            output = root / "output"
            (catalog / "files").mkdir(parents=True)
            workbook_path = catalog / "files" / "fixture.xlsx"
            Workbook().save(workbook_path)

            def write_index(workbook_id: str, digest: str) -> None:
                with (catalog / "INDEX.csv").open(
                    "w", newline="", encoding="utf-8"
                ) as handle:
                    writer = csv.DictWriter(
                        handle,
                        fieldnames=[
                            "id",
                            "category",
                            "extension",
                            "bytes",
                            "sha256",
                            "archive_relative_path",
                        ],
                    )
                    writer.writeheader()
                    writer.writerow(
                        {
                            "id": workbook_id,
                            "category": "hydraulics",
                            "extension": ".xlsx",
                            "bytes": workbook_path.stat().st_size,
                            "sha256": digest,
                            "archive_relative_path": "files/fixture.xlsx",
                        }
                    )

            write_index("..\\..\\escape", "0" * 64)
            with self.assertRaises(ValueError):
                capture_catalog(catalog, output, None, None, True, 30, True)

            write_index("0" * 16, "0" * 64)
            with self.assertRaises(ValueError):
                capture_catalog(catalog, output, None, None, True, 30, True)

            self.assertFalse((output / "workbooks" / ("0" * 16 + ".json")).exists())

    def test_catalog_is_captured_per_workbook_then_stream_merged(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalog = root / "catalog"
            files = catalog / "files"
            output = root / "private-output"
            public = root / "public-output"
            files.mkdir(parents=True)
            workbook_path = files / "fixture.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Hydraulics"
            sheet["A1"] = "Flow (gpm)"
            sheet["A2"] = "Pressure loss (psi)"
            sheet["B2"] = "=B1*2"
            workbook.save(workbook_path)
            digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
            with (catalog / "INDEX.csv").open(
                "w", newline="", encoding="utf-8"
            ) as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "id",
                        "category",
                        "extension",
                        "bytes",
                        "sha256",
                        "original_name",
                        "archive_relative_path",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "id": digest[:16],
                        "category": "hydraulics",
                        "extension": ".xlsx",
                        "bytes": workbook_path.stat().st_size,
                        "sha256": digest,
                        "original_name": "fixture.xlsx",
                        "archive_relative_path": "files/fixture.xlsx",
                    }
                )

            capture_summary = capture_catalog(
                catalog=catalog,
                output_root=output,
                static_reader=None,
                olevba_path=None,
                skip_vba=True,
                timeout_seconds=30,
                force=True,
            )
            merge_summary = merge_captures(output_root=output, public_dir=public)

            capture_files = list((output / "workbooks").glob("*.json"))
            self.assertEqual(capture_summary["captured"], 1)
            self.assertEqual(len(capture_files), 1)
            captured = json.loads(capture_files[0].read_text(encoding="utf-8"))
            family = captured["extraction"]["formula_families"][0]
            self.assertEqual(family["representative_formula"], "=B1*2")
            self.assertEqual(family["occurrence_count"], 1)
            self.assertNotIn("formulas", captured["extraction"])
            self.assertEqual(merge_summary["workbooks"], 1)
            with (public / "CALCULATION_INVENTORY.csv").open(
                newline="", encoding="utf-8"
            ) as handle:
                calculations = list(csv.DictReader(handle))
            self.assertEqual(len(calculations), 1)
            self.assertEqual(calculations[0]["cell_ranges"], "B2")
            self.assertEqual(calculations[0]["layout_kind"], "single-cell")
            self.assertTrue((public / "ANALYSIS_SUMMARY.json").is_file())
            public_text = "\n".join(
                path.read_text(encoding="utf-8") for path in public.glob("*.csv")
            )
            self.assertNotIn("=B1*2", public_text)
            self.assertNotIn(str(root), public_text)


if __name__ == "__main__":
    unittest.main()
