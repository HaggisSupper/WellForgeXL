#!/usr/bin/env python3
"""Statically inventory drilling workbook calculations, units, and VBA.

Modern OOXML workbooks are parsed without Excel. Legacy binary workbooks are
handled by a standalone Rust static reader. VBA is extracted separately with
oletools and is never executed.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


OPENXML_EXTENSIONS = {".xlsx", ".xlsm"}
BINARY_EXTENSIONS = {".xls", ".xlsb"}
# Reserved Office compatibility password used by Microsoft for embedded/protected files.
OFFICE_COMPATIBILITY_PASSWORD = "VelvetSweatshop"
VOLATILE_FUNCTIONS = {
    "CELL",
    "INFO",
    "INDIRECT",
    "NOW",
    "OFFSET",
    "RAND",
    "RANDBETWEEN",
    "TODAY",
}
MAX_INPUT_BYTES = 512 * 1024 * 1024
MAX_DECLARED_CELLS_PER_SHEET = 20_000_000
MAX_FORMULAS_PER_WORKBOOK = 2_000_000
MAX_READER_JSON_BYTES = 512 * 1024 * 1024
MAX_OLEVBA_JSON_BYTES = 512 * 1024 * 1024
WORKBOOK_ID_PATTERN = re.compile(r"^[0-9a-f]{16}$")
SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")


@dataclass(frozen=True)
class UnitSpec:
    canonical_unit: str
    dimension: str
    si_unit: str
    conversion_kind: str
    pattern: re.Pattern[str]


def _unit_spec(
    canonical_unit: str,
    dimension: str,
    si_unit: str,
    pattern: str,
    conversion_kind: str = "scale",
) -> UnitSpec:
    return UnitSpec(
        canonical_unit=canonical_unit,
        dimension=dimension,
        si_unit=si_unit,
        conversion_kind=conversion_kind,
        pattern=re.compile(pattern, re.IGNORECASE),
    )


# Compound units precede their components so overlapping matches resolve to the
# most informative engineering quantity.
UNIT_SPECS: tuple[UnitSpec, ...] = (
    _unit_spec(
        "Btu/(h*ft^2*degF)",
        "heat_transfer_coefficient",
        "W/(m^2*K)",
        r"(?P<unit>BTU\s*/\s*(?:h|hr)\s*[-/]?\s*ft(?:\^?2|²)\s*[-/]?\s*(?:°\s*F|deg\s*F|F))",
    ),
    _unit_spec(
        "W/(m^2*K)",
        "heat_transfer_coefficient",
        "W/(m^2*K)",
        r"(?P<unit>W\s*/\s*(?:m(?:\^?2|²))\s*[-·*]?\s*K)",
    ),
    _unit_spec(
        "Btu/(h*ft*degF)",
        "thermal_conductivity",
        "W/(m*K)",
        r"(?P<unit>BTU\s*/\s*(?:h|hr)\s*[-/]?\s*ft\s*[-/]?\s*(?:°\s*F|deg\s*F|F))",
    ),
    _unit_spec(
        "W/(m*K)",
        "thermal_conductivity",
        "W/(m*K)",
        r"(?P<unit>W\s*/\s*m\s*[-·*]?\s*K)",
    ),
    _unit_spec(
        "kJ/(kg*K)",
        "specific_heat_capacity",
        "J/(kg*K)",
        r"(?P<unit>kJ\s*/\s*\(?\s*kg\s*[-·*]?\s*K\s*\)?)",
    ),
    _unit_spec(
        "J/(kg*K)",
        "specific_heat_capacity",
        "J/(kg*K)",
        r"(?P<unit>J\s*/\s*\(?\s*kg\s*[-·*]?\s*K\s*\)?)",
    ),
    _unit_spec(
        "Btu/(lbm*degF)",
        "specific_heat_capacity",
        "J/(kg*K)",
        r"(?P<unit>BTU\s*/\s*lbm\s*[-/]?\s*(?:°\s*F|deg\s*F|F))",
    ),
    _unit_spec(
        "lbf*s^n/(100*ft^2)",
        "rheology_consistency",
        "Pa*s^n",
        r"(?P<unit>lbf\s*[·*]?\s*s(?:\^?n|ⁿ)\s*/\s*100\s*ft(?:\^?2|²))",
        "exponent-dependent",
    ),
    _unit_spec(
        "Pa*s^n",
        "rheology_consistency",
        "Pa*s^n",
        r"(?P<unit>Pa\s*[·*]?\s*s(?:\^?n|ⁿ))",
        "exponent-dependent",
    ),
    _unit_spec(
        "lbf/(100*ft^2)",
        "yield_stress",
        "Pa",
        r"(?P<unit>lbf\s*/\s*100\s*ft(?:\^?2|²))",
    ),
    _unit_spec(
        "deg/100ft",
        "curvature",
        "rad/m",
        r"(?P<unit>(?:°|deg(?:ree)?s?)\s*/\s*100\s*ft)",
    ),
    _unit_spec(
        "deg/30m",
        "curvature",
        "rad/m",
        r"(?P<unit>(?:°|deg(?:ree)?s?)\s*/\s*30\s*m)",
    ),
    _unit_spec(
        "psi/ft",
        "pressure_gradient",
        "Pa/m",
        r"(?P<unit>psi\s*/\s*ft)",
    ),
    _unit_spec(
        "kPa/m",
        "pressure_gradient",
        "Pa/m",
        r"(?P<unit>kPa\s*/\s*m)",
    ),
    _unit_spec(
        "MPa/m",
        "pressure_gradient",
        "Pa/m",
        r"(?P<unit>MPa\s*/\s*m)",
    ),
    _unit_spec(
        "degF/100ft",
        "temperature_gradient",
        "K/m",
        r"(?P<unit>(?:°\s*F|deg\s*F)\s*/\s*100\s*ft)",
    ),
    _unit_spec(
        "degC/100m",
        "temperature_gradient",
        "K/m",
        r"(?P<unit>(?:°\s*C|deg\s*C)\s*/\s*100\s*m)",
    ),
    _unit_spec(
        "kg/m^3",
        "density",
        "kg/m^3",
        r"(?P<unit>kg\s*/\s*m(?:\^?3|³))",
    ),
    _unit_spec(
        "g/cm^3",
        "density",
        "kg/m^3",
        r"(?P<unit>g\s*/\s*(?:cc|cm(?:\^?3|³)))",
    ),
    _unit_spec(
        "ppg",
        "density",
        "kg/m^3",
        r"(?P<unit>ppg)",
    ),
    _unit_spec(
        "lbm/gal",
        "density",
        "kg/m^3",
        r"(?P<unit>lbm?\s*/\s*(?:US\s*)?gal)",
    ),
    _unit_spec(
        "lbm/ft^3",
        "density",
        "kg/m^3",
        r"(?P<unit>lbm?\s*/\s*ft(?:\^?3|³))",
    ),
    _unit_spec(
        "m^3/s",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>m(?:\^?3|³)\s*/\s*s(?:ec(?:ond)?s?)?)",
    ),
    _unit_spec(
        "m^3/min",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>m(?:\^?3|³)\s*/\s*min(?:ute)?s?)",
    ),
    _unit_spec(
        "L/s",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>l(?:iter|itre)?s?\s*/\s*s(?:ec(?:ond)?s?)?)",
    ),
    _unit_spec(
        "L/min",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>l(?:iter|itre)?s?\s*/\s*min(?:ute)?s?)",
    ),
    _unit_spec(
        "gal/min",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>gpm|(?:US\s*)?gal(?:lon)?s?\s*/\s*min(?:ute)?s?)",
    ),
    _unit_spec(
        "bbl/min",
        "volumetric_flow",
        "m^3/s",
        r"(?P<unit>bpm|bbls?\s*/\s*min(?:ute)?s?)",
    ),
    _unit_spec(
        "m/s",
        "velocity",
        "m/s",
        r"(?P<unit>m\s*/\s*s(?:ec(?:ond)?s?)?)",
    ),
    _unit_spec(
        "ft/s",
        "velocity",
        "m/s",
        r"(?P<unit>ft\s*/\s*s(?:ec(?:ond)?s?)?)",
    ),
    _unit_spec(
        "ft/min",
        "velocity",
        "m/s",
        r"(?P<unit>ft\s*/\s*min(?:ute)?s?)",
    ),
    _unit_spec(
        "m/h",
        "penetration_rate",
        "m/s",
        r"(?P<unit>m\s*/\s*(?:h|hr|hour)s?)",
    ),
    _unit_spec(
        "ft/h",
        "penetration_rate",
        "m/s",
        r"(?P<unit>ft\s*/\s*(?:h|hr|hour)s?)",
    ),
    _unit_spec(
        "kN*m",
        "torque",
        "N*m",
        r"(?P<unit>kN\s*[-·*]\s*m)\b",
    ),
    _unit_spec(
        "N*m",
        "torque",
        "N*m",
        r"(?P<unit>N\s*[-·*]\s*m)\b",
    ),
    _unit_spec(
        "ft*lbf",
        "torque",
        "N*m",
        r"(?P<unit>(?:ft\s*[-·*]\s*lbf|lbf\s*[-·*]\s*ft))\b",
    ),
    _unit_spec(
        "kg/m",
        "linear_mass",
        "kg/m",
        r"(?P<unit>kg\s*/\s*m)\b",
    ),
    _unit_spec(
        "lbm/ft",
        "linear_mass",
        "kg/m",
        r"(?P<unit>lbm?\s*/\s*ft)\b",
    ),
    _unit_spec(
        "mPa*s",
        "dynamic_viscosity",
        "Pa*s",
        r"(?P<unit>mPa\s*[·*]\s*s)\b",
    ),
    _unit_spec(
        "Pa*s",
        "dynamic_viscosity",
        "Pa*s",
        r"(?P<unit>Pa\s*[·*]\s*s)\b",
    ),
    _unit_spec(
        "cP",
        "dynamic_viscosity",
        "Pa*s",
        r"(?P<unit>cP|centipoise)\b",
    ),
    _unit_spec("m^2", "area", "m^2", r"(?P<unit>m(?:\^?2|²))\b"),
    _unit_spec("ft^2", "area", "m^2", r"(?P<unit>ft(?:\^?2|²))\b"),
    _unit_spec("in^2", "area", "m^2", r"(?P<unit>in(?:\^?2|²))\b"),
    _unit_spec("mm^2", "area", "m^2", r"(?P<unit>mm(?:\^?2|²))\b"),
    _unit_spec("m^3", "volume", "m^3", r"(?P<unit>m(?:\^?3|³))\b"),
    _unit_spec("ft^3", "volume", "m^3", r"(?P<unit>ft(?:\^?3|³))\b"),
    _unit_spec("in^3", "volume", "m^3", r"(?P<unit>in(?:\^?3|³))\b"),
    _unit_spec("bbl", "volume", "m^3", r"(?P<unit>bbls?|barrels?)\b"),
    _unit_spec("gal", "volume", "m^3", r"(?P<unit>(?:US\s*)?gals?|gallons?)\b"),
    _unit_spec("kPa", "pressure", "Pa", r"(?P<unit>kPa)\b"),
    _unit_spec("MPa", "pressure", "Pa", r"(?P<unit>MPa)\b"),
    _unit_spec("Pa", "pressure", "Pa", r"(?P<unit>Pa)\b"),
    _unit_spec("ksi", "pressure", "Pa", r"(?P<unit>ksi)\b"),
    _unit_spec("psia", "pressure", "Pa", r"(?P<unit>psia)\b"),
    _unit_spec("psig", "pressure", "Pa", r"(?P<unit>psig)\b"),
    _unit_spec("psi", "pressure", "Pa", r"(?P<unit>psi)\b"),
    _unit_spec("bara", "pressure", "Pa", r"(?P<unit>bara)\b"),
    _unit_spec("barg", "pressure", "Pa", r"(?P<unit>barg)\b"),
    _unit_spec("bar", "pressure", "Pa", r"(?P<unit>bar)\b"),
    _unit_spec("kN", "force", "N", r"(?P<unit>kN)\b"),
    _unit_spec("N", "force", "N", r"(?P<unit>newtons?|(?-i:N))\b"),
    _unit_spec("klbf", "force", "N", r"(?P<unit>klbf|kips?)\b"),
    _unit_spec("lbf", "force", "N", r"(?P<unit>lbf)\b"),
    _unit_spec("kg", "mass", "kg", r"(?P<unit>kg)\b"),
    _unit_spec("lbm", "mass", "kg", r"(?P<unit>lbm)\b"),
    _unit_spec("rad/s", "angular_velocity", "rad/s", r"(?P<unit>rad\s*/\s*s)\b"),
    _unit_spec("rpm", "rotational_speed", "rad/s", r"(?P<unit>rpm)\b"),
    _unit_spec(
        "strokes/min", "pump_rate", "1/s", r"(?P<unit>spm|strokes?\s*/\s*min)\b"
    ),
    _unit_spec("rad", "angle", "rad", r"(?P<unit>radians?|rad)\b"),
    _unit_spec(
        "deg",
        "angle",
        "rad",
        r"(?P<unit>degrees?|deg|°)(?!\s*[CF])\b",
    ),
    _unit_spec(
        "degF",
        "temperature",
        "K",
        r"(?P<unit>°\s*F|deg(?:ree)?s?\s*F|Fahrenheit)\b",
        "affine",
    ),
    _unit_spec(
        "degC",
        "temperature",
        "K",
        r"(?P<unit>°\s*C|deg(?:ree)?s?\s*C|Celsius)\b",
        "affine",
    ),
    _unit_spec("K", "temperature", "K", r"(?P<unit>Kelvin|(?-i:K))\b", "affine"),
    _unit_spec("kW", "power", "W", r"(?P<unit>kW)\b"),
    _unit_spec("W", "power", "W", r"(?P<unit>watts?|(?-i:W))\b"),
    _unit_spec("hp", "power", "W", r"(?P<unit>HHP|hp|horsepower)\b"),
    _unit_spec("kJ", "energy", "J", r"(?P<unit>kJ)\b"),
    _unit_spec("J", "energy", "J", r"(?P<unit>joules?|(?-i:J))\b"),
    _unit_spec("Btu", "energy", "J", r"(?P<unit>BTU)\b"),
    _unit_spec("Hz", "frequency", "1/s", r"(?P<unit>Hz)\b"),
    _unit_spec("ppm", "mass_fraction", "1", r"(?P<unit>ppm)\b"),
    _unit_spec("percent", "fraction", "1", r"(?P<unit>wt\s*%|vol\s*%|%)"),
    _unit_spec("specific-gravity", "density_ratio", "1", r"(?P<unit>\bSG\b)"),
    _unit_spec("mm", "length", "m", r"(?P<unit>mm)\b"),
    _unit_spec("cm", "length", "m", r"(?P<unit>cm)\b"),
    _unit_spec("ft", "length", "m", r"(?P<unit>feet|foot|ft)\b"),
    _unit_spec(
        "in",
        "length",
        "m",
        r"(?:^|[(\[,;/]\s*)(?P<unit>in\.?|inch(?:es)?|\")(?=\s*(?:[\]),;/]|$))",
    ),
    _unit_spec(
        "m",
        "length",
        "m",
        r"(?:^|[(\[,;/]\s*)(?P<unit>m|met(?:er|re)s?)(?=\s*(?:[\]),;/]|$))",
    ),
    _unit_spec("s", "time", "s", r"(?P<unit>seconds?|sec)\b"),
    _unit_spec("min", "time", "s", r"(?P<unit>minutes?|min)\b"),
    _unit_spec("h", "time", "s", r"(?P<unit>hours?|hrs?)\b"),
)


SI_MULTIPLIERS: dict[str, float] = {
    "Btu/(h*ft^2*degF)": 5.6782633411,
    "W/(m^2*K)": 1.0,
    "Btu/(h*ft*degF)": 1.7307346664,
    "W/(m*K)": 1.0,
    "kJ/(kg*K)": 1000.0,
    "J/(kg*K)": 1.0,
    "Btu/(lbm*degF)": 4186.8,
    "Pa*s^n": 1.0,
    "lbf/(100*ft^2)": 0.4788025898,
    "deg/100ft": 0.0005729578,
    "deg/30m": 0.0005817764,
    "psi/ft": 22620.594793,
    "kPa/m": 1000.0,
    "MPa/m": 1_000_000.0,
    "degF/100ft": 0.01822637795,
    "degC/100m": 0.01,
    "kg/m^3": 1.0,
    "g/cm^3": 1000.0,
    "ppg": 119.82642732,
    "lbm/gal": 119.82642732,
    "lbm/ft^3": 16.018463374,
    "m^3/s": 1.0,
    "m^3/min": 1.0 / 60.0,
    "L/s": 0.001,
    "L/min": 0.001 / 60.0,
    "gal/min": 0.003785411784 / 60.0,
    "bbl/min": 0.158987294928 / 60.0,
    "m/s": 1.0,
    "ft/s": 0.3048,
    "ft/min": 0.3048 / 60.0,
    "m/h": 1.0 / 3600.0,
    "ft/h": 0.3048 / 3600.0,
    "kN*m": 1000.0,
    "N*m": 1.0,
    "ft*lbf": 1.3558179483,
    "kg/m": 1.0,
    "lbm/ft": 1.4881639436,
    "mPa*s": 0.001,
    "Pa*s": 1.0,
    "cP": 0.001,
    "m^2": 1.0,
    "ft^2": 0.09290304,
    "in^2": 0.00064516,
    "mm^2": 0.000001,
    "m^3": 1.0,
    "ft^3": 0.028316846592,
    "in^3": 0.000016387064,
    "bbl": 0.158987294928,
    "gal": 0.003785411784,
    "Pa": 1.0,
    "kPa": 1000.0,
    "MPa": 1_000_000.0,
    "ksi": 6_894_757.293168,
    "psi": 6894.757293168,
    "psia": 6894.757293168,
    "psig": 6894.757293168,
    "bar": 100_000.0,
    "bara": 100_000.0,
    "barg": 100_000.0,
    "N": 1.0,
    "kN": 1000.0,
    "lbf": 4.4482216153,
    "klbf": 4448.2216153,
    "kg": 1.0,
    "lbm": 0.45359237,
    "rad/s": 1.0,
    "rpm": 0.1047197551,
    "strokes/min": 1.0 / 60.0,
    "rad": 1.0,
    "deg": 0.01745329252,
    "degC": 1.0,
    "degF": 5.0 / 9.0,
    "K": 1.0,
    "kW": 1000.0,
    "W": 1.0,
    "hp": 745.69987158,
    "kJ": 1000.0,
    "J": 1.0,
    "Btu": 1055.05585262,
    "Hz": 1.0,
    "ppm": 0.000001,
    "percent": 0.01,
    "specific-gravity": 1.0,
    "mm": 0.001,
    "cm": 0.01,
    "ft": 0.3048,
    "in": 0.0254,
    "m": 1.0,
    "s": 1.0,
    "min": 60.0,
    "h": 3600.0,
}
SI_OFFSETS: dict[str, float] = {
    "degC": 273.15,
    "degF": 255.3722222222,
    "K": 0.0,
}


TOPIC_RULES: tuple[tuple[str, re.Pattern[str]], ...] = tuple(
    (topic, re.compile(pattern, re.IGNORECASE))
    for topic, pattern in (
        (
            "hydraulics.fluid-properties",
            r"\b(?:pvt|fluid propert|compressib|density|specific gravity)\b",
        ),
        (
            "hydraulics.rheology",
            r"\b(?:rheolog|bingham|power law|herschel|yield power|plastic viscosity|yield point)\b",
        ),
        ("hydraulics.reynolds-number", r"\b(?:reynolds|reynold|\bre\b)\b"),
        (
            "hydraulics.flow-regime",
            r"\b(?:flow regime|laminar|turbulent|transition(?:al)?)\b",
        ),
        (
            "hydraulics.friction-factor",
            r"\b(?:friction factor|darcy factor|fanning factor)\b",
        ),
        (
            "hydraulics.pressure-loss",
            r"\b(?:pressure (?:loss|drop)|friction loss|parasitic loss|standpipe pressure|\bspp\b)\b",
        ),
        (
            "hydraulics.hydrostatic-pressure",
            r"\b(?:hydrostatic|static pressure|bottomhole pressure|bhp)\b",
        ),
        (
            "hydraulics.equivalent-circulating-density",
            r"\b(?:ecd|equivalent circulating density)\b",
        ),
        (
            "hydraulics.bit-nozzle",
            r"\b(?:nozzle|tfa|total flow area|jet velocity|impact force|hsi|hydraulic horsepower|bit pressure)\b",
        ),
        (
            "hydraulics.annular-velocity",
            r"\b(?:annular velocity|pipe velocity|flow velocity)\b",
        ),
        (
            "hydraulics.hole-cleaning",
            r"\b(?:hole clean|cuttings transport|slip velocity|transport ratio|bed height)\b",
        ),
        ("hydraulics.surge-swab", r"\b(?:surge|swab|tripping speed)\b"),
        ("hydraulics.lag", r"\b(?:lag time|lag volume|strokes to surface)\b"),
        (
            "hydraulics.two-phase-flow",
            r"\b(?:two[ -]?phase|2[ -]?phase|gas liquid|multiphase)\b",
        ),
        (
            "torque-drag.axial-load",
            r"\b(?:axial load|tension|compression|hook ?load|over ?pull|drag)\b",
        ),
        (
            "torque-drag.torque",
            r"\b(?:rotary torque|torque and drag|drag[ &/]+torque|surface torque|bit torque)\b",
        ),
        ("torque-drag.contact-force", r"\b(?:contact force|normal force|side force)\b"),
        ("torque-drag.buckling", r"\b(?:buckl|sinusoidal|helical)\b"),
        (
            "torque-drag.stress-capacity",
            r"\b(?:von mises|triaxial|combined stress|yield strength|safety factor|design factor)\b",
        ),
        (
            "torque-drag.connection-capacity",
            r"\b(?:connection|make[ -]?up torque|torsional yield|joint strength)\b",
        ),
        (
            "directional.trajectory",
            r"\b(?:trajectory|survey calculation|minimum curvature|radius of curvature)\b",
        ),
        (
            "directional.coordinates",
            r"\b(?:northing|easting|tvd|vertical section|departure|closure)\b",
        ),
        ("directional.dogleg-severity", r"\b(?:dogleg|dls|build rate|turn rate)\b"),
        ("directional.toolface", r"\b(?:toolface|tool face|slide orient)\b"),
        (
            "directional.anti-collision",
            r"\b(?:anti[ -]?collision|separation factor|ellipsoid of uncertainty)\b",
        ),
        (
            "bha.geometry",
            r"\b(?:bha|bottom hole assembly|stabilizer|gauge|clearance|bit diameter|tool diameter)\b",
        ),
        (
            "bha.motor-performance",
            r"\b(?:mud motor|pdm|power section|motor rpm|motor torque|rotor|stator)\b",
        ),
        (
            "bha.bending-loading",
            r"\b(?:bending|flex shaft|flexi shaft|side force|neutral point)\b",
        ),
        ("bha.tool-life", r"\b(?:tool life|durability|wear|swelling|fatigue)\b"),
        (
            "bha.casing-clearance",
            r"\b(?:casing jam|jam risk|casing clearance|pass through)\b",
        ),
        (
            "casing.capacity",
            r"\b(?:casing capacity|pipe capacity|displacement|internal volume)\b",
        ),
        ("casing.strength", r"\b(?:burst|collapse|tension rating|triaxial)\b"),
        (
            "cementing.volume",
            r"\b(?:cement volume|annular volume|slurry volume|cement yield|excess)\b",
        ),
        (
            "cementing.hydrostatic",
            r"\b(?:cement hydrostatic|placement pressure|slurry density|mix water)\b",
        ),
        (
            "well-control.kill-density",
            r"\b(?:kill mud|kill weight|kill density|wait and weight|driller.?s method)\b",
        ),
        (
            "well-control.kick-tolerance",
            r"\b(?:kick tolerance|influx volume|kick volume)\b",
        ),
        (
            "well-control.pressure-schedule",
            r"\b(?:icp|fcp|initial circulating pressure|final circulating pressure|choke schedule)\b",
        ),
        (
            "well-control.maasp",
            r"\b(?:maasp|maximum allowable annular surface pressure)\b",
        ),
        (
            "well-control.gas-behavior",
            r"\b(?:gas migration|gas expansion|boyle|influx)\b",
        ),
        (
            "thermal.geothermal-gradient",
            r"\b(?:geothermal|formation temperature|temperature gradient|static temperature)\b",
        ),
        (
            "thermal.heat-exchange",
            r"\b(?:heat exchang\w*|heat transfer|overall u|\bua\b|thermal conduct\w*|convection|conduction)\b",
        ),
        (
            "thermal.reverse-circulation",
            r"\b(?:reverse circulation|reverse flow|counter[ -]?current|counterflow)\b",
        ),
        (
            "thermal.transient",
            r"\b(?:transient temperature|time dependent|thermal diffusiv|thermal inertia)\b",
        ),
        (
            "thermal.tool-limits",
            r"\b(?:thermal limit|temperature limit|derat|electronics temperature)\b",
        ),
        (
            "cross-cutting.unit-conversion",
            r"\b(?:unit conversion|convert units?|metric|imperial|oilfield units?)\b",
        ),
        (
            "cross-cutting.lookup",
            r"\b(?:lookup|table|database|catalog|material propert)\b",
        ),
        (
            "cross-cutting.optimization",
            r"\b(?:solver|goal seek|optimi[sz]|sensitivity)\b",
        ),
        (
            "cross-cutting.validation",
            r"\b(?:validation|verification|benchmark|test case|error check)\b",
        ),
    )
)


RISK_RULES: tuple[tuple[str, re.Pattern[str]], ...] = tuple(
    (signal, re.compile(pattern, re.IGNORECASE | re.MULTILINE))
    for signal, pattern in (
        (
            "external-process",
            r"\b(?:Shell|WScript\.Shell|ShellExecute|cmd\.exe|powershell(?:\.exe)?)\b",
        ),
        ("com-automation", r"\b(?:CreateObject|GetObject)\s*\("),
        (
            "filesystem-write",
            r"\b(?:FileCopy|Kill|MkDir|RmDir|SaveAs|Open\s+.+\s+For\s+(?:Output|Append|Binary)|FileSystemObject)\b",
        ),
        (
            "network-access",
            r"\b(?:XMLHTTP|WinHttp|URLDownloadToFile|FollowHyperlink|https?://|ftp://)\b",
        ),
        ("database-access", r"\b(?:ADODB|DAO\.|OLEDB|ODBC)\b"),
        ("native-api", r"^\s*(?:Public|Private)?\s*Declare\b|\bPtrSafe\b|\bLib\s+\""),
        ("environment-access", r"\b(?:Environ\$?|GetSetting|SaveSetting|RegOpenKey)\b"),
        ("code-mutation", r"\b(?:VBProject|VBComponents|CodeModule|AddFromString)\b"),
    )
)


PROCEDURE_START = re.compile(
    r"^\s*(?:(Public|Private|Friend|Static)\s+)?"
    r"(Sub|Function|Property\s+(?:Get|Let|Set))\s+"
    r"([A-Za-z_][A-Za-z0-9_]*)\b",
    re.IGNORECASE,
)
PROCEDURE_END = re.compile(r"^\s*End\s+(?:Sub|Function|Property)\b", re.IGNORECASE)
FORMULA_FUNCTION = re.compile(r"(?<![A-Z0-9_.])(?:_XLFN\.)?([A-Z][A-Z0-9_.]*)\s*\(")
FORMULA_STRING = re.compile(r'"(?:[^"]|"")*"')
EXTERNAL_BOOK = re.compile(
    r"\[[^\]\r\n]+\](?=(?:'(?:[^']|'')*'|[A-Z0-9_. ]+)!)", re.IGNORECASE
)
SHEET_REFERENCE = re.compile(r"(?:'(?:[^']|'')+'|[A-Z_][A-Z0-9_. ]*)!", re.IGNORECASE)
A1_REFERENCE = re.compile(
    r"(?<![A-Z0-9_.])\$?[A-Z]{1,3}\$?[1-9][0-9]*(?::\$?[A-Z]{1,3}\$?[1-9][0-9]*)?(?![A-Z0-9_.(])",
    re.IGNORECASE,
)
R1C1_REFERENCE = re.compile(
    r"(?<![A-Z0-9_.])R(?:\[-?\d+\]|\d+)?C(?:\[-?\d+\]|\d+)?(?![A-Z0-9_.(])",
    re.IGNORECASE,
)
PUBLIC_EXCEL_FUNCTIONS = {
    "ABS",
    "ACOS",
    "AND",
    "ASIN",
    "ATAN",
    "ATAN2",
    "AVERAGE",
    "AVERAGEIF",
    "AVERAGEIFS",
    "CEILING",
    "CHOOSE",
    "COLUMN",
    "COLUMNS",
    "CONCAT",
    "CONCATENATE",
    "COS",
    "COUNT",
    "COUNTA",
    "COUNTIF",
    "COUNTIFS",
    "DEGREES",
    "EXP",
    "FALSE",
    "FIND",
    "FLOOR",
    "HLOOKUP",
    "IF",
    "IFERROR",
    "IFNA",
    "IFS",
    "INDEX",
    "INDIRECT",
    "ISBLANK",
    "ISERROR",
    "ISNA",
    "LEFT",
    "LEN",
    "LN",
    "LOG",
    "LOG10",
    "LOOKUP",
    "MATCH",
    "MAX",
    "MID",
    "MIN",
    "MOD",
    "NA",
    "NOT",
    "NOW",
    "OFFSET",
    "OR",
    "PI",
    "POWER",
    "RADIANS",
    "RAND",
    "RANDBETWEEN",
    "RIGHT",
    "ROUND",
    "ROUNDDOWN",
    "ROUNDUP",
    "ROW",
    "ROWS",
    "SIGN",
    "SIN",
    "SQRT",
    "SUM",
    "SUMIF",
    "SUMIFS",
    "TAN",
    "TEXT",
    "TODAY",
    "TRIM",
    "TRUE",
    "VLOOKUP",
    "XLOOKUP",
    "XMATCH",
}


def normalize_formula(formula: str) -> str:
    """Return a stable, non-reversible structural formula signature."""

    normalized = str(formula).strip().upper()
    normalized = FORMULA_STRING.sub("<TEXT>", normalized)
    normalized = EXTERNAL_BOOK.sub("[BOOK]", normalized)
    normalized = SHEET_REFERENCE.sub("SHEET!", normalized)
    normalized = A1_REFERENCE.sub("<REF>", normalized)
    normalized = R1C1_REFERENCE.sub("<REF>", normalized)
    return re.sub(r"\s+", "", normalized)


def extract_formula_functions(formula: str) -> list[str]:
    """List unique worksheet function names in order of first appearance."""

    seen: set[str] = set()
    functions: list[str] = []
    for match in FORMULA_FUNCTION.finditer(str(formula).upper()):
        name = match.group(1)
        if name not in seen:
            seen.add(name)
            functions.append(name)
    return functions


def formula_has_external_reference(formula: str) -> bool:
    """Distinguish external workbook links from structured table references."""

    return bool(EXTERNAL_BOOK.search(str(formula)))


def public_formula_functions(functions: Iterable[str]) -> list[str]:
    """Keep standard Excel names while masking workbook-specific UDF names."""

    public: set[str] = set()
    for function in functions:
        normalized = str(function).upper().removeprefix("_XLFN.")
        public.add(normalized if normalized in PUBLIC_EXCEL_FUNCTIONS else "UDF")
    return sorted(public)


def _unit_semantics(canonical_unit: str, dimension: str) -> dict[str, Any]:
    pressure_basis = {
        "psia": "absolute",
        "bara": "absolute",
        "psig": "gauge",
        "barg": "gauge",
    }.get(
        canonical_unit, "unspecified" if dimension == "pressure" else "not-applicable"
    )
    temperature_kind = (
        "absolute"
        if dimension == "temperature"
        else "difference-per-length"
        if dimension == "temperature_gradient"
        else "not-applicable"
    )
    reference_state = (
        "context-required"
        if dimension in {"volume", "volumetric_flow"}
        else "not-applicable"
    )
    quantity_kind = {
        "ppg": "mass-density-or-mud-weight",
        "lbm/gal": "mass-density-or-mud-weight",
    }.get(canonical_unit, dimension)
    return {
        "native_unit": canonical_unit,
        "si_multiplier": SI_MULTIPLIERS.get(canonical_unit, ""),
        "si_offset": SI_OFFSETS.get(canonical_unit, 0.0)
        if dimension == "temperature"
        else 0.0,
        "pressure_basis": pressure_basis,
        "temperature_kind": temperature_kind,
        "reference_state": reference_state,
        "quantity_kind": quantity_kind,
    }


def detect_units(text: Any, source_kind: str = "cell-text") -> list[dict[str, Any]]:
    """Detect explicit engineering-unit tokens while suppressing overlaps."""

    if text is None:
        return []
    value = str(text).replace("\u00a0", " ")
    occupied: list[tuple[int, int]] = []
    found: list[dict[str, str]] = []
    for spec in UNIT_SPECS:
        for match in spec.pattern.finditer(value):
            span = match.span("unit") if "unit" in match.groupdict() else match.span()
            if any(max(span[0], start) < min(span[1], end) for start, end in occupied):
                continue
            native_text = match.group("unit").strip()
            if native_text in {"N", "W", "J", "K"}:
                context_terms = {
                    "N": r"\b(?:force|load|weight|wob|thrust|tension|compression)\b",
                    "W": r"\b(?:power|watt|heat|thermal)\b",
                    "J": r"\b(?:energy|work|joule|heat)\b",
                    "K": r"\b(?:temperature|temp|kelvin|thermal)\b",
                }
                contextual_label = bool(
                    re.search(context_terms[native_text], value, re.IGNORECASE)
                )
                if not contextual_label:
                    continue
            occupied.append(span)
            found.append(
                {
                    "native_text": native_text,
                    "canonical_unit": spec.canonical_unit,
                    "dimension": spec.dimension,
                    "si_unit": spec.si_unit,
                    "conversion_kind": spec.conversion_kind,
                    "source_kind": source_kind,
                    **_unit_semantics(spec.canonical_unit, spec.dimension),
                }
            )
    return found


def classify_calculation_topics(text: Any) -> list[str]:
    """Return all matching calculation topics rather than one folder label."""

    value = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", str(text or ""))
    value = value.replace("_", " ")
    return [topic for topic, pattern in TOPIC_RULES if pattern.search(value)]


def _executable_lines(code: str) -> list[str]:
    lines: list[str] = []
    for line in code.splitlines():
        stripped = line.strip()
        if (
            not stripped
            or stripped.startswith("'")
            or re.match(r"(?i)^Rem(?:\s|$)", stripped)
        ):
            continue
        lines.append(line)
    return lines


def _execution_trigger(name: str) -> str:
    folded = name.casefold()
    if folded in {"auto_open", "auto_close", "auto_exec", "autoexit"}:
        return "auto-exec"
    if folded.startswith("workbook_") or folded in {"document_open", "document_close"}:
        return "workbook-event"
    if folded.startswith("worksheet_"):
        return "worksheet-event"
    if folded.startswith("userform_"):
        return "userform-event"
    return "explicit-call"


def _risk_signals(code: str) -> list[str]:
    executable = "\n".join(_executable_lines(code))
    return [signal for signal, pattern in RISK_RULES if pattern.search(executable)]


def analyze_vba_modules(modules: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Inventory VBA procedures and behavior signals without executing source."""

    rows: list[dict[str, Any]] = []
    for module in modules:
        module_name = str(module.get("module_name") or module.get("vba_filename") or "")
        code = str(module.get("code") or "")
        lines = code.splitlines()
        index = 0
        while index < len(lines):
            start = PROCEDURE_START.match(lines[index])
            if not start:
                index += 1
                continue
            visibility = (start.group(1) or "Public").lower()
            kind = re.sub(r"\s+", "-", start.group(2).lower())
            name = start.group(3)
            end = index + 1
            while end < len(lines) and not PROCEDURE_END.match(lines[end]):
                end += 1
            if end < len(lines):
                end += 1
            body = "\n".join(lines[index:end])
            executable_count = len(_executable_lines(body))
            rows.append(
                {
                    "module_name": module_name,
                    "procedure_name": name,
                    "procedure_kind": kind,
                    "visibility": visibility,
                    "start_line": index + 1,
                    "end_line": end,
                    "executable_lines": executable_count,
                    "execution_trigger": _execution_trigger(name),
                    "calculation_topics": classify_calculation_topics(
                        name + "\n" + body
                    ),
                    "units": sorted(
                        {
                            item["canonical_unit"]
                            for item in detect_units(body, source_kind="vba-source")
                        }
                    ),
                    "risk_signals": _risk_signals(body),
                }
            )
            index = max(end, index + 1)
    return rows


def _formula_text(value: Any) -> tuple[str | None, str]:
    if isinstance(value, str) and value.startswith("="):
        return value, "regular"
    text = getattr(value, "text", None)
    if isinstance(text, str) and text.startswith("="):
        name = type(value).__name__.casefold()
        kind = (
            "array"
            if "array" in name
            else "data-table"
            if "datatable" in name
            else name
        )
        return text, kind
    return None, ""


def _nearby_labels(
    labels: dict[tuple[int, int], str], row: int, column: int, limit: int = 8
) -> list[str]:
    candidates: list[tuple[int, int, str]] = []
    for label_row in range(max(1, row - 4), row + 5):
        for label_column in range(max(1, column - 4), column + 5):
            row_delta = abs(label_row - row)
            column_delta = abs(label_column - column)
            same_row = label_row == row and column_delta <= 4
            same_column = label_column == column and row_delta <= 4
            nearby = row_delta <= 2 and column_delta <= 2
            text = labels.get((label_row, label_column))
            if text is not None and (same_row or same_column or nearby):
                direction_bias = 0 if label_column <= column and label_row <= row else 1
                candidates.append((row_delta + column_delta, direction_bias, text))
    candidates.sort(key=lambda item: (item[0], item[1], item[2].casefold()))
    result: list[str] = []
    seen: set[str] = set()
    for _, _, text in candidates:
        compact = re.sub(r"\s+", " ", text).strip()[:240]
        if compact and compact not in seen:
            seen.add(compact)
            result.append(compact)
        if len(result) >= limit:
            break
    return result


def _append_unit_mentions(
    target: list[dict[str, Any]],
    text: Any,
    source_kind: str,
    sheet_name: str,
    cell: str,
) -> None:
    for mention in detect_units(text, source_kind=source_kind):
        target.append({"sheet_name": sheet_name, "cell": cell, **mention})


def extract_ooxml_workbook(path: Path) -> dict[str, Any]:
    """Extract formulas, labels, names, and units from an OOXML workbook."""

    from openpyxl import load_workbook

    input_bytes = path.stat().st_size
    if input_bytes > MAX_INPUT_BYTES:
        raise ValueError(
            f"Workbook exceeds {MAX_INPUT_BYTES} byte static-analysis limit"
        )
    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    result: dict[str, Any] = {
        "status": "ok",
        "method": "openpyxl-read-only",
        "sheets": [],
        "formulas": [],
        "defined_names": [],
        "unit_mentions": [],
        "external_link_count": len(getattr(workbook, "_external_links", [])),
        "calculation": {},
    }
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        for attribute in (
            "calcMode",
            "iterate",
            "iterateCount",
            "iterateDelta",
            "fullCalcOnLoad",
            "forceFullCalc",
        ):
            result["calculation"][attribute] = getattr(calculation, attribute, None)

    try:
        for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]
            visibility = str(getattr(sheet, "sheet_state", "visible"))
            if not hasattr(sheet, "iter_rows"):
                result["sheets"].append(
                    {
                        "sheet_index": sheet_index,
                        "sheet_name": sheet_name,
                        "sheet_kind": "chart",
                        "visibility": visibility,
                        "rows_used": 0,
                        "columns_used": 0,
                        "populated_cells": 0,
                        "formula_cells": 0,
                    }
                )
                continue

            declared_rows = int(getattr(sheet, "max_row", 0) or 0)
            declared_columns = int(getattr(sheet, "max_column", 0) or 0)
            if declared_rows * declared_columns > MAX_DECLARED_CELLS_PER_SHEET:
                raise ValueError(
                    f"Sheet {sheet_index} declares more than "
                    f"{MAX_DECLARED_CELLS_PER_SHEET} cells"
                )

            labels: dict[tuple[int, int], str] = {}
            pending_formulas: list[dict[str, Any]] = []
            populated_cells = 0
            maximum_row = 0
            maximum_column = 0
            for row_cells in sheet.iter_rows():
                for cell in row_cells:
                    value = cell.value
                    if value is None:
                        continue
                    populated_cells += 1
                    maximum_row = max(maximum_row, cell.row)
                    maximum_column = max(maximum_column, cell.column)
                    formula, formula_kind = _formula_text(value)
                    number_format = str(getattr(cell, "number_format", "") or "")
                    if formula is not None:
                        pending_formulas.append(
                            {
                                "sheet_name": sheet_name,
                                "cell": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "formula": formula,
                                "formula_kind": formula_kind,
                                "number_format": number_format,
                            }
                        )
                        if (
                            len(result["formulas"]) + len(pending_formulas)
                            > MAX_FORMULAS_PER_WORKBOOK
                        ):
                            raise ValueError(
                                f"Workbook exceeds {MAX_FORMULAS_PER_WORKBOOK} formula limit"
                            )
                    elif isinstance(value, str):
                        labels[(cell.row, cell.column)] = value
                        _append_unit_mentions(
                            result["unit_mentions"],
                            value,
                            "cell-text",
                            sheet_name,
                            cell.coordinate,
                        )
                    if number_format and number_format.casefold() != "general":
                        _append_unit_mentions(
                            result["unit_mentions"],
                            number_format,
                            "number-format",
                            sheet_name,
                            cell.coordinate,
                        )

            workbook_context = f"{path.stem} {sheet_name}"
            for formula_row in pending_formulas:
                context_labels = _nearby_labels(
                    labels, formula_row["row"], formula_row["column"]
                )
                context = " | ".join([workbook_context, *context_labels])
                formula = formula_row["formula"]
                functions = extract_formula_functions(formula)
                normalized = normalize_formula(formula)
                context_units = detect_units(context, source_kind="formula-context")
                format_units = detect_units(
                    formula_row["number_format"], source_kind="number-format"
                )
                result["formulas"].append(
                    {
                        **formula_row,
                        "normalized_formula": normalized,
                        "family_id": hashlib.sha256(
                            normalized.encode("utf-8")
                        ).hexdigest()[:16],
                        "functions": functions,
                        "context_labels": context_labels,
                        "calculation_topics": classify_calculation_topics(context),
                        "units": sorted(
                            {
                                item["canonical_unit"]
                                for item in [*context_units, *format_units]
                            }
                        ),
                        "external_reference": formula_has_external_reference(formula),
                        "volatile_functions": sorted(
                            set(functions) & VOLATILE_FUNCTIONS
                        ),
                    }
                )

            result["sheets"].append(
                {
                    "sheet_index": sheet_index,
                    "sheet_name": sheet_name,
                    "sheet_kind": "worksheet",
                    "visibility": visibility,
                    "rows_used": maximum_row,
                    "columns_used": maximum_column,
                    "populated_cells": populated_cells,
                    "formula_cells": len(pending_formulas),
                }
            )

        for definition in workbook.defined_names.values():
            refers_to = str(getattr(definition, "attr_text", "") or "")
            result["defined_names"].append(
                {
                    "name": str(getattr(definition, "name", "") or ""),
                    "scope_sheet_index": getattr(definition, "localSheetId", None),
                    "hidden": bool(getattr(definition, "hidden", False)),
                    "refers_to": refers_to,
                    "reference_kind": "external"
                    if formula_has_external_reference(refers_to)
                    else "formula"
                    if refers_to.startswith("=") and "!" not in refers_to
                    else "range",
                    "functions": extract_formula_functions(refers_to),
                    "calculation_topics": classify_calculation_topics(
                        f"{getattr(definition, 'name', '')} {refers_to}"
                    ),
                    "units": sorted(
                        {
                            item["canonical_unit"]
                            for item in detect_units(
                                f"{getattr(definition, 'name', '')} {refers_to}",
                                source_kind="defined-name",
                            )
                        }
                    ),
                }
            )
    finally:
        workbook.close()
    return result


CELL_REFERENCE = re.compile(r"^\$?([A-Z]{1,3})\$?([1-9][0-9]*)$", re.IGNORECASE)


def _cell_coordinates(reference: str) -> tuple[int, int]:
    match = CELL_REFERENCE.match(str(reference))
    if not match:
        raise ValueError(f"Invalid cell reference: {reference}")
    column = 0
    for character in match.group(1).upper():
        column = column * 26 + ord(character) - ord("A") + 1
    return int(match.group(2)), column


def convert_calamine_audit(
    payload: dict[str, Any], workbook_stem: str
) -> dict[str, Any]:
    """Convert one Rust static-reader document into the common audit model."""

    result: dict[str, Any] = {
        "status": "ok",
        "method": "calamine-static",
        "sheets": [],
        "formulas": [],
        "defined_names": [],
        "unit_mentions": [],
        "external_link_count": None,
        "calculation": {},
    }
    for sheet in payload.get("sheets", []):
        sheet_name = str(sheet.get("sheet_name", ""))
        labels: dict[tuple[int, int], str] = {}
        for text_value in sheet.get("text_cells", []):
            text = str(text_value.get("text", ""))
            occurrences = max(1, int(text_value.get("occurrences", 1)))
            cells = [str(cell) for cell in text_value.get("cells", [])]
            for cell in cells:
                try:
                    labels[_cell_coordinates(cell)] = text
                except ValueError:
                    continue
            representative_cell = cells[0] if cells else ""
            for mention in detect_units(text, source_kind="cell-text"):
                result["unit_mentions"].append(
                    {
                        "sheet_name": sheet_name,
                        "cell": representative_cell,
                        "occurrences": occurrences,
                        **mention,
                    }
                )

        pending_formulas: list[dict[str, Any]] = []
        for formula_item in sheet.get("formulas", []):
            cell = str(formula_item.get("cell", ""))
            formula = str(formula_item.get("formula", ""))
            parsed = bool(formula_item.get("parsed", True))
            if formula and not formula.startswith("="):
                formula = f"={formula}"
            try:
                row, column = _cell_coordinates(cell)
            except ValueError:
                row, column = 0, 0
            pending_formulas.append(
                {
                    "sheet_name": sheet_name,
                    "cell": cell,
                    "row": row,
                    "column": column,
                    "formula": formula,
                    "formula_kind": (
                        "unparsed"
                        if not parsed
                        else "xlm-macro"
                        if sheet.get("sheet_kind") == "macro-sheet"
                        else "regular"
                    ),
                    "number_format": "",
                }
            )

        workbook_context = f"{workbook_stem} {sheet_name}"
        for formula_row in pending_formulas:
            context_labels = _nearby_labels(
                labels, formula_row["row"], formula_row["column"]
            )
            context = " | ".join([workbook_context, *context_labels])
            formula = formula_row["formula"]
            functions = extract_formula_functions(formula)
            normalized = normalize_formula(formula)
            result["formulas"].append(
                {
                    **formula_row,
                    "normalized_formula": normalized,
                    "family_id": hashlib.sha256(normalized.encode("utf-8")).hexdigest()[
                        :16
                    ],
                    "functions": functions,
                    "context_labels": context_labels,
                    "calculation_topics": classify_calculation_topics(context),
                    "units": sorted(
                        {
                            item["canonical_unit"]
                            for item in detect_units(
                                context, source_kind="formula-context"
                            )
                        }
                    ),
                    "external_reference": formula_has_external_reference(formula),
                    "volatile_functions": sorted(set(functions) & VOLATILE_FUNCTIONS),
                }
            )

        result["sheets"].append(
            {
                "sheet_index": int(sheet.get("sheet_index", len(result["sheets"]) + 1)),
                "sheet_name": sheet_name,
                "sheet_kind": str(sheet.get("sheet_kind", "worksheet")),
                "visibility": str(sheet.get("visibility", "visible")),
                "rows_used": int(sheet.get("rows_used", 0)),
                "columns_used": int(sheet.get("columns_used", 0)),
                "populated_cells": int(sheet.get("populated_cells", 0)),
                "formula_cells": len(pending_formulas),
                "warnings": [str(item) for item in sheet.get("warnings", [])],
            }
        )

    for definition in payload.get("defined_names", []):
        name = str(definition.get("name", ""))
        refers_to = str(definition.get("formula", ""))
        result["defined_names"].append(
            {
                "name": name,
                "scope_sheet_index": None,
                "hidden": None,
                "metadata_status": "scope-and-visibility-unsupported-by-calamine",
                "refers_to": refers_to,
                "reference_kind": "external"
                if formula_has_external_reference(refers_to)
                else "formula"
                if refers_to.startswith("=") and "!" not in refers_to
                else "range",
                "functions": extract_formula_functions(refers_to),
                "calculation_topics": classify_calculation_topics(
                    f"{name} {refers_to}"
                ),
                "units": sorted(
                    {
                        item["canonical_unit"]
                        for item in detect_units(
                            f"{name} {refers_to}", source_kind="defined-name"
                        )
                    }
                ),
            }
        )
    return result


def parse_olevba_payload(payload: Any) -> dict[str, Any]:
    """Reduce one olevba JSON payload while retaining source only in local data."""

    documents = payload if isinstance(payload, list) else [payload]
    modules: list[dict[str, Any]] = []
    pcode_records: list[dict[str, Any]] = []
    indicators: list[dict[str, str]] = []
    messages: list[dict[str, str]] = []
    for document in documents:
        if not isinstance(document, dict):
            continue
        if document.get("type") == "msg":
            messages.append(
                {
                    "level": str(document.get("level", "")),
                    "message": str(document.get("msg", "")),
                }
            )
        for indicator in document.get("analysis") or []:
            if not isinstance(indicator, dict):
                continue
            indicators.append(
                {
                    "type": str(indicator.get("type", "")),
                    "keyword": str(indicator.get("keyword", "")),
                    "description": str(indicator.get("description", "")),
                }
            )
        for module in document.get("macros") or []:
            if not isinstance(module, dict):
                continue
            module_name = str(module.get("vba_filename", ""))
            code = module.get("code")
            code_text = "" if code is None else str(code)
            if module_name.casefold() == "vba_p-code.txt":
                pcode_records.append(
                    {
                        "record_name": "vba-pcode-disassembly",
                        "line_count": len(code_text.splitlines()),
                        "code": code_text,
                    }
                )
                continue
            suffix = Path(module_name).suffix.casefold()
            module_kind = {
                ".bas": "standard-module",
                ".cls": "class-or-document-module",
                ".frm": "user-form",
            }.get(suffix, "unknown-module")
            modules.append(
                {
                    "module_name": module_name,
                    "module_kind": module_kind,
                    "ole_stream": str(module.get("ole_stream", "")),
                    "source_status": "available" if code is not None else "unavailable",
                    "line_count": len(code_text.splitlines()),
                    "code": code_text,
                }
            )
    procedures = analyze_vba_modules(modules)
    procedure_counts = Counter(row["module_name"] for row in procedures)
    for module in modules:
        module["procedure_count"] = procedure_counts[module["module_name"]]
    return {
        "status": "ok",
        "has_vba": bool(modules or pcode_records),
        "module_count": len(modules),
        "pcode_record_count": len(pcode_records),
        "procedure_count": len(procedures),
        "event_procedure_count": sum(
            row["execution_trigger"] != "explicit-call" for row in procedures
        ),
        "modules": modules,
        "pcode_records": pcode_records,
        "procedures": procedures,
        "indicators": indicators,
        "messages": messages,
    }


def _sheet_id(sheet_index: int) -> str:
    return f"s{sheet_index:04d}"


def _column_letters(column: int) -> str:
    letters: list[str] = []
    value = column
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def _format_cell(row: int, column: int) -> str:
    return f"{_column_letters(column)}{row}"


def _compress_positions(positions: Iterable[tuple[int, int]]) -> tuple[list[str], str]:
    cells = sorted(set(positions))
    if not cells:
        return [], "unknown"
    rows = sorted({row for row, _ in cells})
    columns = sorted({column for _, column in cells})
    if len(cells) == 1:
        row, column = cells[0]
        return [_format_cell(row, column)], "single-cell"
    if len(cells) == len(rows) * len(columns):
        start = _format_cell(rows[0], columns[0])
        end = _format_cell(rows[-1], columns[-1])
        if len(columns) == 1:
            return [f"{start}:{end}"], "filled-column"
        if len(rows) == 1:
            return [f"{start}:{end}"], "filled-row"
        return [f"{start}:{end}"], "rectangular-table"

    ranges: list[str] = []
    by_column: dict[int, list[int]] = defaultdict(list)
    for row, column in cells:
        by_column[column].append(row)
    for column, column_rows in sorted(by_column.items()):
        run_start = column_rows[0]
        run_end = column_rows[0]
        for row in column_rows[1:]:
            if row == run_end + 1:
                run_end = row
                continue
            start = _format_cell(run_start, column)
            end = _format_cell(run_end, column)
            ranges.append(start if run_start == run_end else f"{start}:{end}")
            run_start = run_end = row
        start = _format_cell(run_start, column)
        end = _format_cell(run_end, column)
        ranges.append(start if run_start == run_end else f"{start}:{end}")
    return ranges, "multi-area"


def compact_formula_families(analysis: dict[str, Any]) -> dict[str, Any]:
    """Replace cell-by-cell formula records with compact structural families."""

    if "formulas" not in analysis:
        return analysis
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for formula in analysis.get("formulas", []):
        sheet_name = str(formula.get("sheet_name", ""))
        family_id = str(formula.get("family_id", ""))
        formula_kind = str(formula.get("formula_kind", "regular"))
        key = (sheet_name, family_id, formula_kind)
        group = grouped.setdefault(
            key,
            {
                "sheet_name": sheet_name,
                "family_id": family_id,
                "formula_kind": formula_kind,
                "representative_formula": str(formula.get("formula", "")),
                "normalized_formula": str(formula.get("normalized_formula", "")),
                "occurrence_count": 0,
                "positions": [],
                "functions": set(),
                "volatile_functions": set(),
                "calculation_topics": set(),
                "units": set(),
                "context_labels": [],
                "external_reference_count": 0,
            },
        )
        try:
            group["positions"].append(_cell_coordinates(str(formula.get("cell", ""))))
        except ValueError:
            pass
        group["occurrence_count"] += 1
        group["functions"].update(formula.get("functions", []))
        group["volatile_functions"].update(formula.get("volatile_functions", []))
        group["calculation_topics"].update(formula.get("calculation_topics", []))
        group["units"].update(formula.get("units", []))
        for label in formula.get("context_labels", []):
            if (
                label not in group["context_labels"]
                and len(group["context_labels"]) < 12
            ):
                group["context_labels"].append(label)
        group["external_reference_count"] += int(
            bool(formula.get("external_reference", False))
        )

    sheet_order = {
        str(sheet.get("sheet_name", "")): int(sheet.get("sheet_index", 0))
        for sheet in analysis.get("sheets", [])
    }
    families: list[dict[str, Any]] = []
    for group in grouped.values():
        ranges, layout_kind = _compress_positions(group.pop("positions"))
        families.append(
            {
                **group,
                "cell_ranges": ranges,
                "layout_kind": layout_kind,
                "functions": sorted(group["functions"]),
                "volatile_functions": sorted(group["volatile_functions"]),
                "calculation_topics": sorted(group["calculation_topics"]),
                "units": sorted(group["units"]),
            }
        )
    families.sort(
        key=lambda family: (
            sheet_order.get(family["sheet_name"], 0),
            family["family_id"],
            family["formula_kind"],
        )
    )
    compacted = dict(analysis)
    compacted.pop("formulas", None)
    compacted["formula_families"] = families
    return compacted


def build_calculation_rows(
    metadata: dict[str, Any], analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    """Build privacy-safe formula-family rows for one workbook."""

    workbook_id = str(metadata.get("id", ""))
    sheets = {
        str(sheet.get("sheet_name", "")): sheet for sheet in analysis.get("sheets", [])
    }
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    source_formulas = analysis.get("formulas")
    if source_formulas is None:
        source_formulas = []
        for family in analysis.get("formula_families", []):
            grouped[
                (
                    str(family.get("sheet_name", "")),
                    str(family.get("family_id", "")),
                    str(family.get("formula_kind", "regular")),
                )
            ] = {
                "occurrence_count": int(family.get("occurrence_count", 0)),
                "functions": set(family.get("functions", [])),
                "volatile_functions": set(family.get("volatile_functions", [])),
                "calculation_topics": set(family.get("calculation_topics", [])),
                "units": set(family.get("units", [])),
                "external_reference_count": int(
                    family.get("external_reference_count", 0)
                ),
                "layout_kind": str(family.get("layout_kind", "")),
                "cell_range_count": len(family.get("cell_ranges", [])),
                "cell_ranges": [str(item) for item in family.get("cell_ranges", [])],
                "positions": [],
            }
    for formula in source_formulas:
        sheet_name = str(formula.get("sheet_name", ""))
        family_id = str(formula.get("family_id", ""))
        formula_kind = str(formula.get("formula_kind", "regular"))
        key = (sheet_name, family_id, formula_kind)
        group = grouped.setdefault(
            key,
            {
                "occurrence_count": 0,
                "functions": set(),
                "volatile_functions": set(),
                "calculation_topics": set(),
                "units": set(),
                "external_reference_count": 0,
                "layout_kind": "",
                "cell_range_count": 0,
                "cell_ranges": [],
                "positions": [],
            },
        )
        group["occurrence_count"] += 1
        group["functions"].update(formula.get("functions", []))
        group["volatile_functions"].update(formula.get("volatile_functions", []))
        group["calculation_topics"].update(formula.get("calculation_topics", []))
        group["units"].update(formula.get("units", []))
        group["external_reference_count"] += int(
            bool(formula.get("external_reference", False))
        )
        group["positions"].append(
            (int(formula.get("row", 0)), int(formula.get("column", 0)))
        )

    fallback_topic = {
        "hydraulics": "hydraulics.unresolved",
        "torque-drag-drillstring": "torque-drag.unresolved",
        "directional": "directional.unresolved",
        "bha-tools": "bha.unresolved",
        "cementing-casing": "casing-cementing.unresolved",
        "well-control": "well-control.unresolved",
        "thermal": "thermal.unresolved",
        "general-drilling": "general-drilling.unresolved",
        "uncategorized": "unresolved",
    }.get(str(metadata.get("category", "")), "unresolved")
    rows: list[dict[str, Any]] = []
    for family_index, (
        (sheet_name, _private_family_id, formula_kind),
        group,
    ) in enumerate(sorted(grouped.items()), start=1):
        sheet = sheets.get(sheet_name, {})
        sheet_index = int(sheet.get("sheet_index", 0))
        topics = sorted(group["calculation_topics"])
        cell_ranges = group["cell_ranges"]
        layout_kind = group["layout_kind"]
        if group["positions"]:
            cell_ranges, layout_kind = _compress_positions(group["positions"])
        rows.append(
            {
                "workbook_id": workbook_id,
                "category": str(metadata.get("category", "")),
                "extension": str(metadata.get("extension", "")),
                "sheet_index": sheet_index,
                "sheet_id": _sheet_id(sheet_index),
                "sheet_kind": str(sheet.get("sheet_kind", "worksheet")),
                "formula_family_id": f"f{family_index:06d}",
                "formula_kind": formula_kind,
                "implementation": "xlm-macro-formula"
                if sheet.get("sheet_kind") == "macro-sheet"
                else "worksheet-formula",
                "occurrence_count": group["occurrence_count"],
                "layout_kind": layout_kind,
                "cell_range_count": len(cell_ranges),
                "cell_ranges": "|".join(cell_ranges),
                "functions": "|".join(public_formula_functions(group["functions"])),
                "volatile_functions": "|".join(sorted(group["volatile_functions"])),
                "external_reference_count": group["external_reference_count"],
                "calculation_topics": "|".join(topics or [fallback_topic]),
                "classification_evidence": "local-context"
                if topics
                else "catalog-category",
                "detected_units": "|".join(sorted(group["units"])),
            }
        )
    return rows


FORBIDDEN_PUBLIC_FIELDS = {
    "archive_relative_path",
    "code",
    "formula",
    "module_name",
    "normalized_formula",
    "original_name",
    "procedure_name",
    "raw_formula",
    "refers_to",
    "sheet_name",
    "source_path",
    "vba_source",
}
ABSOLUTE_PATH_PATTERN = re.compile(r"(?:[A-Za-z]:[\\/]|\\\\[^\\]+\\)")
POSIX_PATH_PATTERN = re.compile(r"(?:^|\s)/(?:[^/\s]+/)+[^/\s]+")
FILE_URI_PATTERN = re.compile(r"\bfile://", re.IGNORECASE)
EMAIL_PATTERN = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
SPREADSHEET_FORMULA_PREFIX = re.compile(r"^[ \v\f]*[=+\-@\t\r]")


def assert_public_records_safe(records: Iterable[dict[str, Any]]) -> None:
    """Reject private workbook content from rows intended for Git."""

    for record in records:
        forbidden = FORBIDDEN_PUBLIC_FIELDS.intersection(record)
        if forbidden:
            raise ValueError(
                f"Public record contains forbidden fields: {sorted(forbidden)}"
            )
        for key, value in record.items():
            if not isinstance(value, str):
                continue
            if (
                ABSOLUTE_PATH_PATTERN.search(value)
                or POSIX_PATH_PATTERN.search(value)
                or FILE_URI_PATTERN.search(value)
            ):
                raise ValueError(f"Public field {key!r} contains an absolute path")
            if EMAIL_PATTERN.search(value):
                raise ValueError(f"Public field {key!r} contains an email address")
            if SPREADSHEET_FORMULA_PREFIX.search(value):
                raise ValueError(f"Public field {key!r} contains a CSV formula prefix")


def _hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
    os.replace(temporary, path)


def _empty_extraction(method: str, error: str) -> dict[str, Any]:
    return {
        "status": "error",
        "method": method,
        "error": error,
        "sheets": [],
        "formulas": [],
        "defined_names": [],
        "unit_mentions": [],
        "external_link_count": None,
        "calculation": {},
    }


def _load_json_bounded(path: Path, byte_limit: int) -> Any:
    size = path.stat().st_size
    if size > byte_limit:
        raise ValueError(f"JSON output exceeds {byte_limit} byte limit")
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def _read_tail(path: Path, byte_limit: int = 4000) -> str:
    with path.open("rb") as handle:
        handle.seek(max(0, path.stat().st_size - byte_limit))
        return handle.read(byte_limit).decode("utf-8", errors="replace")


def validate_calamine_payload(payload: Any) -> dict[str, Any]:
    """Validate the Rust reader contract before trusting its extraction status."""

    if not isinstance(payload, dict) or payload.get("schema_version") != "1.0.0":
        raise ValueError("Unsupported or missing Calamine audit schema version")
    if not isinstance(payload.get("extension"), str):
        raise ValueError("Calamine audit extension must be a string")
    sheets = payload.get("sheets")
    names = payload.get("defined_names")
    if not isinstance(sheets, list) or not isinstance(names, list):
        raise ValueError("Calamine audit must contain sheet and defined-name arrays")
    for sheet in sheets:
        if not isinstance(sheet, dict):
            raise ValueError("Calamine sheet entry must be an object")
        for field in ("sheet_index", "rows_used", "columns_used", "populated_cells"):
            value = sheet.get(field)
            if not isinstance(value, int) or value < 0:
                raise ValueError(f"Calamine sheet field {field} must be non-negative")
        for field in ("sheet_name", "sheet_kind", "visibility"):
            if not isinstance(sheet.get(field), str):
                raise ValueError(f"Calamine sheet field {field} must be a string")
        if not isinstance(sheet.get("text_cells"), list) or not isinstance(
            sheet.get("formulas"), list
        ):
            raise ValueError("Calamine sheet text_cells and formulas must be arrays")
        if (
            int(sheet["rows_used"]) * int(sheet["columns_used"])
            > MAX_DECLARED_CELLS_PER_SHEET
        ):
            raise ValueError("Calamine sheet exceeds declared cell limit")
        for formula in sheet["formulas"]:
            if not isinstance(formula, dict) or not isinstance(
                formula.get("cell"), str
            ):
                raise ValueError("Calamine formula entry has an invalid cell")
            if not isinstance(formula.get("formula"), str):
                raise ValueError("Calamine formula entry has an invalid expression")
    if sum(len(sheet["formulas"]) for sheet in sheets) > MAX_FORMULAS_PER_WORKBOOK:
        raise ValueError("Calamine audit exceeds formula limit")
    for definition in names:
        if not isinstance(definition, dict) or not all(
            isinstance(definition.get(field), str) for field in ("name", "formula")
        ):
            raise ValueError("Calamine defined-name entry is invalid")
    return payload


def _run_openxml_reader(
    path: Path,
    staging_directory: Path,
    timeout_seconds: int,
) -> dict[str, Any]:
    """Parse and compact one OOXML workbook in a disposable subprocess."""

    staging_directory.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=".openxml.json", dir=staging_directory, delete=False
    ) as output_handle:
        output_path = Path(output_handle.name)
    with tempfile.NamedTemporaryFile(
        suffix=".openxml.stderr", dir=staging_directory, delete=False
    ) as error_handle:
        error_path = Path(error_handle.name)
    try:
        with error_path.open("wb") as error_output:
            completed = subprocess.run(
                [
                    sys.executable,
                    str(Path(__file__).resolve()),
                    "--internal-openxml-input",
                    str(path),
                    "--internal-openxml-output",
                    str(output_path),
                ],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=error_output,
                timeout=timeout_seconds,
            )
        if completed.returncode != 0:
            raise RuntimeError(
                f"OOXML reader exited {completed.returncode}: {_read_tail(error_path).strip()}"
            )
        payload = _load_json_bounded(output_path, MAX_READER_JSON_BYTES)
        if not isinstance(payload, dict) or payload.get("status") != "ok":
            raise ValueError("OOXML subprocess returned an invalid extraction document")
        return payload
    finally:
        output_path.unlink(missing_ok=True)
        error_path.unlink(missing_ok=True)


def _run_calamine_reader(
    path: Path,
    reader: Path,
    staging_directory: Path,
    timeout_seconds: int,
    workbook_stem: str | None = None,
) -> dict[str, Any]:
    staging_directory.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=".json", dir=staging_directory, delete=False
    ) as handle:
        temporary_path = Path(handle.name)
    with tempfile.NamedTemporaryFile(
        suffix=".calamine.stderr", dir=staging_directory, delete=False
    ) as handle:
        error_path = Path(handle.name)
    try:
        with error_path.open("wb") as error_output:
            completed = subprocess.run(
                [
                    str(reader),
                    "--input",
                    str(path),
                    "--output",
                    str(temporary_path),
                ],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=error_output,
                timeout=timeout_seconds,
            )
        if completed.returncode != 0:
            raise RuntimeError(
                f"static reader exited {completed.returncode}: {_read_tail(error_path).strip()}"
            )
        payload = validate_calamine_payload(
            _load_json_bounded(temporary_path, MAX_READER_JSON_BYTES)
        )
        return convert_calamine_audit(payload, workbook_stem or path.stem)
    finally:
        temporary_path.unlink(missing_ok=True)
        error_path.unlink(missing_ok=True)


def _run_calamine_after_standard_decryption(
    path: Path,
    reader: Path,
    decryptor: Path,
    staging_directory: Path,
    timeout_seconds: int,
) -> dict[str, Any]:
    """Decrypt an Office compatibility-protected workbook, then parse it statically."""

    staging_directory.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=path.suffix, dir=staging_directory, delete=False
    ) as handle:
        decrypted_path = Path(handle.name)
    with tempfile.NamedTemporaryFile(
        suffix=".decrypt.stderr", dir=staging_directory, delete=False
    ) as handle:
        error_path = Path(handle.name)
    try:
        with error_path.open("wb") as error_output:
            completed = subprocess.run(
                [
                    str(decryptor),
                    "--password",
                    OFFICE_COMPATIBILITY_PASSWORD,
                    str(path),
                    str(decrypted_path),
                ],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=error_output,
                timeout=timeout_seconds,
            )
        if completed.returncode != 0:
            raise RuntimeError(
                f"static decryptor exited {completed.returncode}: {_read_tail(error_path).strip()}"
            )
        if decrypted_path.stat().st_size > MAX_INPUT_BYTES:
            raise ValueError("Decrypted workbook exceeds static-analysis byte limit")
        result = _run_calamine_reader(
            decrypted_path,
            reader,
            staging_directory,
            timeout_seconds,
            workbook_stem=path.stem,
        )
        result["method"] = "calamine-static-after-standard-office-decryption"
        return result
    finally:
        decrypted_path.unlink(missing_ok=True)
        error_path.unlink(missing_ok=True)


def extract_workbook_static(
    path: Path,
    static_reader: Path | None,
    decryptor: Path | None,
    staging_directory: Path,
    timeout_seconds: int,
) -> dict[str, Any]:
    """Extract one workbook through deterministic, non-executing parser fallbacks."""

    extension = path.suffix.casefold()
    if extension in OPENXML_EXTENSIONS:
        try:
            return _run_openxml_reader(path, staging_directory, timeout_seconds)
        except Exception as primary_error:
            if static_reader is None:
                raise
            try:
                result = _run_calamine_reader(
                    path, static_reader, staging_directory, timeout_seconds
                )
            except RuntimeError as secondary_error:
                error_text = f"{primary_error}; {secondary_error}"
                is_protected = bool(
                    re.search(
                        r"password protected|encrypted", error_text, re.IGNORECASE
                    )
                )
                if not is_protected or decryptor is None:
                    raise
                result = _run_calamine_after_standard_decryption(
                    path,
                    static_reader,
                    decryptor,
                    staging_directory,
                    timeout_seconds,
                )
                result["fallback_reason"] = error_text
                return result
            result["method"] = "calamine-static-ooxml-fallback"
            result["fallback_reason"] = (
                f"{type(primary_error).__name__}: {primary_error}"
            )
            return result

    if extension in BINARY_EXTENSIONS:
        if static_reader is None:
            raise RuntimeError(
                "calamine static reader is required for binary workbooks"
            )
        try:
            return _run_calamine_reader(
                path, static_reader, staging_directory, timeout_seconds
            )
        except RuntimeError as primary_error:
            error_text = str(primary_error)
            is_protected = bool(
                re.search(r"password protected|encrypted", error_text, re.IGNORECASE)
            )
            if not is_protected or decryptor is None:
                raise
            result = _run_calamine_after_standard_decryption(
                path,
                static_reader,
                decryptor,
                staging_directory,
                timeout_seconds,
            )
            result["fallback_reason"] = error_text
            return result

    raise ValueError(f"Unsupported workbook extension: {extension}")


def _decode_olevba_json(path: Path) -> Any:
    if path.stat().st_size > MAX_OLEVBA_JSON_BYTES:
        raise ValueError(f"olevba JSON exceeds {MAX_OLEVBA_JSON_BYTES} byte limit")
    with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
        text = handle.read()
    start = text.find("[")
    end = text.rfind("]")
    if start < 0 or end < start:
        raise ValueError("olevba did not emit a JSON array")
    return json.loads(text[start : end + 1])


def _run_olevba(
    path: Path,
    executable: Path | None,
    staging_directory: Path,
    timeout_seconds: int,
) -> dict[str, Any]:
    if executable is None:
        return {
            "status": "unavailable",
            "pcode_inspection": "not-run",
            "has_vba": None,
            "module_count": 0,
            "procedure_count": 0,
            "event_procedure_count": 0,
            "modules": [],
            "procedures": [],
            "indicators": [],
            "messages": [],
        }
    staging_directory.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=".olevba.json", dir=staging_directory, delete=False
    ) as handle:
        temporary_path = Path(handle.name)
    with tempfile.NamedTemporaryFile(
        suffix=".olevba.stderr", dir=staging_directory, delete=False
    ) as handle:
        error_path = Path(handle.name)
    environment = os.environ.copy()
    environment["PYTHONIOENCODING"] = "utf-8"
    environment["PYTHONDONTWRITEBYTECODE"] = "1"
    try:
        with temporary_path.open("wb") as output, error_path.open("wb") as error_output:
            completed = subprocess.run(
                [str(executable), "-j", str(path)],
                check=False,
                stdout=output,
                stderr=error_output,
                timeout=timeout_seconds,
                env=environment,
            )
        payload = _decode_olevba_json(temporary_path)
        result = parse_olevba_payload(payload)
        result["exit_code"] = completed.returncode
        result["stderr"] = _read_tail(error_path)
        result["pcode_inspection"] = "enabled"
        if completed.returncode != 0:
            result["status"] = "partial"
        return result
    except subprocess.TimeoutExpired:
        return {
            "status": "timeout",
            "pcode_inspection": "not-run",
            "has_vba": None,
            "module_count": 0,
            "procedure_count": 0,
            "event_procedure_count": 0,
            "modules": [],
            "procedures": [],
            "indicators": [],
            "messages": [],
        }
    except (OSError, ValueError, json.JSONDecodeError) as error:
        return {
            "status": "error",
            "pcode_inspection": "not-run",
            "error": f"{type(error).__name__}: {error}",
            "has_vba": None,
            "module_count": 0,
            "procedure_count": 0,
            "event_procedure_count": 0,
            "modules": [],
            "procedures": [],
            "indicators": [],
            "messages": [],
        }
    finally:
        temporary_path.unlink(missing_ok=True)
        error_path.unlink(missing_ok=True)


def _load_catalog_records(catalog: Path) -> list[dict[str, str]]:
    index_path = catalog / "PRIVATE_INDEX.csv"
    if not index_path.is_file():
        index_path = catalog / "INDEX.csv"
    with index_path.open(newline="", encoding="utf-8-sig") as handle:
        records = list(csv.DictReader(handle))
    ids = [str(record.get("id", "")).casefold() for record in records]
    if any(not WORKBOOK_ID_PATTERN.fullmatch(workbook_id) for workbook_id in ids):
        raise ValueError(f"Catalog {index_path.name} contains an invalid workbook id")
    if len(ids) != len(set(ids)):
        raise ValueError(f"Catalog {index_path.name} contains duplicate workbook ids")
    for record, workbook_id in zip(records, ids, strict=True):
        record["id"] = workbook_id
        digest = str(record.get("sha256", "")).casefold()
        if not SHA256_PATTERN.fullmatch(digest) or not digest.startswith(workbook_id):
            raise ValueError(f"Catalog digest is invalid for workbook {workbook_id}")
    return sorted(records, key=lambda row: row["id"])


def capture_catalog(
    catalog: Path,
    output_root: Path,
    static_reader: Path | None,
    olevba_path: Path | None,
    skip_vba: bool,
    timeout_seconds: int,
    force: bool,
    limit: int | None = None,
    workbook_ids: set[str] | None = None,
    msoffcrypto_path: Path | None = None,
) -> dict[str, int]:
    """Capture one independent private JSON document per workbook."""

    catalog = catalog.resolve()
    output_root = output_root.resolve()
    records = _load_catalog_records(catalog)
    if workbook_ids:
        known_ids = {record["id"] for record in records}
        unknown_ids = sorted(workbook_ids - known_ids)
        if unknown_ids:
            raise ValueError(f"Unknown workbook ids: {', '.join(unknown_ids)}")
        records = [record for record in records if record["id"] in workbook_ids]
    if limit is not None:
        records = records[:limit]

    workbook_directory = output_root / "workbooks"
    staging_directory = output_root / "staging"
    workbook_directory.mkdir(parents=True, exist_ok=True)
    staging_directory.mkdir(parents=True, exist_ok=True)
    summary = {"catalog_records": len(records), "captured": 0, "reused": 0, "errors": 0}

    for position, metadata in enumerate(records, start=1):
        workbook_id = metadata["id"]
        destination = (workbook_directory / f"{workbook_id}.json").resolve()
        try:
            destination.relative_to(workbook_directory.resolve())
        except ValueError as error:
            raise ValueError(
                f"Workbook id escapes capture directory: {workbook_id}"
            ) from error
        if destination.exists() and not force:
            summary["reused"] += 1
            print(f"[{position}/{len(records)}] reuse {workbook_id}", flush=True)
            continue

        relative_path = Path(metadata["archive_relative_path"])
        source = (catalog / relative_path).resolve()
        try:
            source.relative_to(catalog)
        except ValueError as error:
            raise ValueError(
                f"Archive path escapes catalog: {relative_path}"
            ) from error

        print(f"[{position}/{len(records)}] capture {workbook_id}", flush=True)
        expected_hash = str(metadata.get("sha256", ""))
        if source.stat().st_size > MAX_INPUT_BYTES:
            raise ValueError(
                f"Workbook {workbook_id} exceeds the static-analysis byte limit"
            )
        computed_hash = _hash_file(source)
        if computed_hash.casefold() != expected_hash.casefold():
            destination.unlink(missing_ok=True)
            raise ValueError(
                f"Workbook {workbook_id} no longer matches its catalog SHA-256"
            )
        try:
            extraction = extract_workbook_static(
                source,
                static_reader,
                msoffcrypto_path,
                staging_directory,
                timeout_seconds,
            )
            extraction = compact_formula_families(extraction)
        except Exception as error:  # One bad workbook must not abort the collection.
            extraction = _empty_extraction(
                "openpyxl-read-only"
                if source.suffix.casefold() in OPENXML_EXTENSIONS
                else "calamine-static",
                f"{type(error).__name__}: {error}",
            )

        if skip_vba:
            macro_analysis = {
                "status": "skipped",
                "pcode_inspection": "not-run",
                "has_vba": None,
                "module_count": 0,
                "procedure_count": 0,
                "event_procedure_count": 0,
                "modules": [],
                "procedures": [],
                "indicators": [],
                "messages": [],
            }
        else:
            macro_analysis = _run_olevba(
                source, olevba_path, staging_directory, timeout_seconds
            )

        document = {
            "analysis_schema_version": "1.0.0",
            "workbook": dict(metadata),
            "integrity": {
                "expected_sha256": expected_hash,
                "computed_sha256": computed_hash,
                "matches_index": bool(expected_hash)
                and computed_hash.casefold() == expected_hash.casefold(),
            },
            "extraction": extraction,
            "macro_analysis": macro_analysis,
        }
        _write_json_atomic(destination, document)
        summary["captured"] += 1
        if extraction.get("status") != "ok" or macro_analysis.get("status") != "ok":
            if not (skip_vba and macro_analysis.get("status") == "skipped"):
                summary["errors"] += 1

        # Deliberately release each workbook document before opening the next.
        del document, extraction, macro_analysis
        import gc

        gc.collect()

    return summary


def _fallback_topic(category: str) -> str:
    return {
        "hydraulics": "hydraulics.unresolved",
        "torque-drag-drillstring": "torque-drag.unresolved",
        "directional": "directional.unresolved",
        "bha-tools": "bha.unresolved",
        "cementing-casing": "casing-cementing.unresolved",
        "well-control": "well-control.unresolved",
        "thermal": "thermal.unresolved",
        "general-drilling": "general-drilling.unresolved",
        "uncategorized": "unresolved",
    }.get(category, "unresolved")


def build_sheet_rows(
    metadata: dict[str, Any], analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    workbook_id = str(metadata.get("id", ""))
    formulas_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for formula in analysis.get("formulas", []):
        formulas_by_sheet[str(formula.get("sheet_name", ""))].append(formula)
    families_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for family in analysis.get("formula_families", []):
        families_by_sheet[str(family.get("sheet_name", ""))].append(family)
    units_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for mention in analysis.get("unit_mentions", []):
        units_by_sheet[str(mention.get("sheet_name", ""))].append(mention)

    rows: list[dict[str, Any]] = []
    for sheet in analysis.get("sheets", []):
        sheet_name = str(sheet.get("sheet_name", ""))
        sheet_formulas = formulas_by_sheet[sheet_name]
        sheet_families = families_by_sheet[sheet_name]
        sheet_units = units_by_sheet[sheet_name]
        topics = {
            topic
            for formula in sheet_formulas
            for topic in formula.get("calculation_topics", [])
        }
        topics.update(
            topic
            for family in sheet_families
            for topic in family.get("calculation_topics", [])
        )
        if not topics:
            topics.update(
                classify_calculation_topics(
                    f"{metadata.get('original_name', '')} {sheet_name}"
                )
            )
        rows.append(
            {
                "workbook_id": workbook_id,
                "category": str(metadata.get("category", "")),
                "extension": str(metadata.get("extension", "")),
                "sheet_index": int(sheet.get("sheet_index", 0)),
                "sheet_id": _sheet_id(int(sheet.get("sheet_index", 0))),
                "sheet_kind": str(sheet.get("sheet_kind", "worksheet")),
                "visibility": str(sheet.get("visibility", "visible")),
                "rows_used": int(sheet.get("rows_used", 0)),
                "columns_used": int(sheet.get("columns_used", 0)),
                "populated_cells": int(sheet.get("populated_cells", 0)),
                "formula_cells": len(sheet_formulas)
                if sheet_formulas
                else sum(
                    int(family.get("occurrence_count", 0)) for family in sheet_families
                ),
                "formula_families": len(sheet_families)
                if sheet_families
                else len({formula.get("family_id", "") for formula in sheet_formulas}),
                "external_reference_formulas": sum(
                    bool(formula.get("external_reference", False))
                    for formula in sheet_formulas
                )
                if sheet_formulas
                else sum(
                    int(family.get("external_reference_count", 0))
                    for family in sheet_families
                ),
                "volatile_formula_cells": sum(
                    bool(formula.get("volatile_functions", []))
                    for formula in sheet_formulas
                )
                if sheet_formulas
                else sum(
                    int(family.get("occurrence_count", 0))
                    for family in sheet_families
                    if family.get("volatile_functions", [])
                ),
                "functions": "|".join(
                    public_formula_functions(
                        function
                        for formula in [*sheet_formulas, *sheet_families]
                        for function in formula.get("functions", [])
                    )
                ),
                "calculation_topics": "|".join(
                    sorted(
                        topics or {_fallback_topic(str(metadata.get("category", "")))}
                    )
                ),
                "unit_evidence_count": sum(
                    int(mention.get("occurrences", 1)) for mention in sheet_units
                ),
                "unit_dimensions": "|".join(
                    sorted(
                        {str(mention.get("dimension", "")) for mention in sheet_units}
                    )
                ),
                "canonical_units": "|".join(
                    sorted(
                        {
                            str(mention.get("canonical_unit", ""))
                            for mention in sheet_units
                        }
                    )
                ),
                "warning_count": len(sheet.get("warnings", [])),
                "extraction_method": str(analysis.get("method", "")),
                "extraction_status": str(analysis.get("status", "")),
            }
        )
    return rows


def _unit_hazard(dimension: str, conversion_kind: str) -> str:
    if conversion_kind == "affine":
        return "absolute-temperature-vs-temperature-difference"
    return {
        "pressure": "gauge-vs-absolute-and-pressure-vs-stress",
        "pressure_gradient": "pressure-gradient-vs-equivalent-density",
        "density": "mass-density-vs-mud-weight-gradient",
        "rheology_consistency": "conversion-depends-on-flow-index",
        "yield_stress": "yield-point-vs-stress-convention",
        "torque": "torque-vs-energy",
        "curvature": "interval-basis-and-angle-convention",
        "volume": "actual-vs-standard-reference-state",
        "volumetric_flow": "actual-vs-standard-reference-state",
        "heat_transfer_coefficient": "area-basis-vs-linear-conductance",
        "thermal_conductivity": "conductivity-vs-overall-coefficient",
    }.get(dimension, "")


def build_unit_rows(
    metadata: dict[str, Any], analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    sheets = {
        str(sheet.get("sheet_name", "")): sheet for sheet in analysis.get("sheets", [])
    }
    grouped: Counter[tuple[Any, ...]] = Counter()
    for mention in analysis.get("unit_mentions", []):
        canonical = str(mention.get("canonical_unit", ""))
        dimension = str(mention.get("dimension", ""))
        semantics = {**_unit_semantics(canonical, dimension), **mention}
        key = (
            str(mention.get("sheet_name", "")),
            str(semantics.get("native_unit", canonical)),
            canonical,
            dimension,
            str(mention.get("si_unit", "")),
            str(mention.get("conversion_kind", "")),
            semantics.get("si_multiplier", ""),
            semantics.get("si_offset", ""),
            str(semantics.get("pressure_basis", "")),
            str(semantics.get("temperature_kind", "")),
            str(semantics.get("reference_state", "")),
            str(semantics.get("quantity_kind", "")),
            str(mention.get("source_kind", "")),
        )
        grouped[key] += int(mention.get("occurrences", 1))
    rows: list[dict[str, Any]] = []
    workbook_id = str(metadata.get("id", ""))
    for key, occurrences in sorted(grouped.items()):
        (
            sheet_name,
            native_unit,
            canonical,
            dimension,
            si_unit,
            conversion_kind,
            si_multiplier,
            si_offset,
            pressure_basis,
            temperature_kind,
            reference_state,
            quantity_kind,
            source_kind,
        ) = key
        sheet = sheets.get(sheet_name, {})
        sheet_index = int(sheet.get("sheet_index", 0))
        rows.append(
            {
                "workbook_id": workbook_id,
                "category": str(metadata.get("category", "")),
                "sheet_index": sheet_index,
                "sheet_id": _sheet_id(sheet_index),
                "observed_unit": native_unit,
                "canonical_unit": canonical,
                "dimension": dimension,
                "canonical_si_unit": si_unit,
                "conversion_kind": conversion_kind,
                "si_multiplier": si_multiplier,
                "si_offset": si_offset,
                "pressure_basis": pressure_basis,
                "temperature_kind": temperature_kind,
                "reference_state": reference_state,
                "quantity_kind": quantity_kind,
                "evidence_source": source_kind,
                "evidence_confidence": "high"
                if source_kind == "cell-text"
                else "medium",
                "occurrences": occurrences,
                "normalization_hazard": _unit_hazard(dimension, conversion_kind),
            }
        )
    return rows


def build_defined_name_rows(
    metadata: dict[str, Any], analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    workbook_id = str(metadata.get("id", ""))
    rows = []
    definitions = sorted(
        analysis.get("defined_names", []),
        key=lambda definition: (
            definition.get("scope_sheet_index") is None,
            int(definition.get("scope_sheet_index") or 0),
            str(definition.get("name", "")).casefold(),
            str(definition.get("refers_to", "")),
        ),
    )
    for name_index, definition in enumerate(definitions, start=1):
        scope = definition.get("scope_sheet_index")
        hidden = definition.get("hidden")
        rows.append(
            {
                "workbook_id": workbook_id,
                "category": str(metadata.get("category", "")),
                "defined_name_id": f"n{name_index:06d}",
                "scope_sheet_index": "" if scope is None else int(scope) + 1,
                "hidden": "" if hidden is None else bool(hidden),
                "metadata_status": str(
                    definition.get("metadata_status", "scope-and-visibility-available")
                ),
                "reference_kind": str(definition.get("reference_kind", "")),
                "external_reference": definition.get("reference_kind") == "external",
                "functions": "|".join(
                    public_formula_functions(definition.get("functions", []))
                ),
                "calculation_topics": "|".join(
                    sorted(definition.get("calculation_topics", []))
                ),
                "detected_units": "|".join(sorted(definition.get("units", []))),
            }
        )
    return rows


def build_vba_rows(
    metadata: dict[str, Any], macro_analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    workbook_id = str(metadata.get("id", ""))
    procedures_by_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for procedure in macro_analysis.get("procedures", []):
        procedures_by_module[str(procedure.get("module_name", ""))].append(procedure)
    rows: list[dict[str, Any]] = []
    modules = sorted(
        macro_analysis.get("modules", []),
        key=lambda module: str(module.get("module_name", "")).casefold(),
    )
    for module_index, module in enumerate(modules, start=1):
        module_name = str(module.get("module_name", ""))
        module_id = f"m{module_index:04d}"
        module_code = str(module.get("code", ""))
        module_topics = classify_calculation_topics(module_code)
        module_units = sorted(
            {
                mention["canonical_unit"]
                for mention in detect_units(module_code, source_kind="vba-source")
            }
        )
        rows.append(
            {
                "workbook_id": workbook_id,
                "category": str(metadata.get("category", "")),
                "module_id": module_id,
                "record_id": module_id,
                "record_kind": "module",
                "module_kind": str(module.get("module_kind", "")),
                "procedure_kind": "",
                "visibility": "",
                "execution_trigger": "",
                "source_status": str(module.get("source_status", "")),
                "source_lines": int(module.get("line_count", 0)),
                "executable_lines": 0,
                "calculation_topics": "|".join(sorted(module_topics)),
                "detected_units": "|".join(module_units),
                "risk_signals": "|".join(_risk_signals(module_code)),
            }
        )
        procedures = sorted(
            procedures_by_module[module_name],
            key=lambda procedure: (
                int(procedure.get("start_line", 0)),
                str(procedure.get("procedure_name", "")).casefold(),
            ),
        )
        for procedure_index, procedure in enumerate(procedures, start=1):
            record_id = f"{module_id}-p{procedure_index:04d}"
            rows.append(
                {
                    "workbook_id": workbook_id,
                    "category": str(metadata.get("category", "")),
                    "module_id": module_id,
                    "record_id": record_id,
                    "record_kind": "procedure",
                    "module_kind": str(module.get("module_kind", "")),
                    "procedure_kind": str(procedure.get("procedure_kind", "")),
                    "visibility": str(procedure.get("visibility", "")),
                    "execution_trigger": str(procedure.get("execution_trigger", "")),
                    "source_status": str(module.get("source_status", "")),
                    "source_lines": int(procedure.get("end_line", 0))
                    - int(procedure.get("start_line", 0))
                    + 1,
                    "executable_lines": int(procedure.get("executable_lines", 0)),
                    "calculation_topics": "|".join(
                        sorted(procedure.get("calculation_topics", []))
                    ),
                    "detected_units": "|".join(sorted(procedure.get("units", []))),
                    "risk_signals": "|".join(sorted(procedure.get("risk_signals", []))),
                }
            )
    return rows


def _indicator_risk_class(indicator: dict[str, Any]) -> str:
    text = f"{indicator.get('keyword', '')} {indicator.get('description', '')}"
    signals = _risk_signals(text)
    if signals:
        return "|".join(signals)
    folded = text.casefold()
    if "auto" in str(indicator.get("type", "")).casefold():
        return "auto-execution"
    if "hex" in folded or "base64" in folded or "obfuscat" in folded:
        return "obfuscation-indicator"
    if str(indicator.get("type", "")).casefold() == "ioc":
        return "indicator-of-compromise"
    return "review-required"


def build_macro_indicator_rows(
    metadata: dict[str, Any], macro_analysis: dict[str, Any]
) -> list[dict[str, Any]]:
    workbook_id = str(metadata.get("id", ""))
    grouped: Counter[tuple[str, str]] = Counter()
    for indicator in macro_analysis.get("indicators", []):
        indicator_type = str(indicator.get("type", ""))
        risk_class = _indicator_risk_class(indicator)
        grouped[(indicator_type, risk_class)] += 1
    return [
        {
            "workbook_id": workbook_id,
            "category": str(metadata.get("category", "")),
            "indicator_type": indicator_type,
            "risk_class": risk_class,
            "occurrences": count,
        }
        for (indicator_type, risk_class), count in sorted(grouped.items())
    ]


WORKBOOK_AUDIT_FIELDS = [
    "workbook_id",
    "category",
    "extension",
    "bytes",
    "integrity_matches_index",
    "extraction_method",
    "extraction_status",
    "static_coverage",
    "sheet_count",
    "hidden_sheet_count",
    "very_hidden_sheet_count",
    "macro_sheet_count",
    "formula_cells",
    "formula_families",
    "defined_names",
    "unit_dimensions",
    "canonical_units",
    "calculation_topics",
    "external_reference_formulas",
    "volatile_formula_cells",
    "iterative_calculation",
    "vba_status",
    "vba_pcode_inspection",
    "has_vba",
    "vba_modules",
    "vba_procedures",
    "event_procedures",
    "macro_indicator_count",
]
SHEET_INVENTORY_FIELDS = [
    "workbook_id",
    "category",
    "extension",
    "sheet_index",
    "sheet_id",
    "sheet_kind",
    "visibility",
    "rows_used",
    "columns_used",
    "populated_cells",
    "formula_cells",
    "formula_families",
    "external_reference_formulas",
    "volatile_formula_cells",
    "functions",
    "calculation_topics",
    "unit_evidence_count",
    "unit_dimensions",
    "canonical_units",
    "warning_count",
    "extraction_method",
    "extraction_status",
]
CALCULATION_INVENTORY_FIELDS = [
    "workbook_id",
    "category",
    "extension",
    "sheet_index",
    "sheet_id",
    "sheet_kind",
    "formula_family_id",
    "formula_kind",
    "implementation",
    "occurrence_count",
    "layout_kind",
    "cell_range_count",
    "cell_ranges",
    "functions",
    "volatile_functions",
    "external_reference_count",
    "calculation_topics",
    "classification_evidence",
    "detected_units",
]
UNIT_INVENTORY_FIELDS = [
    "workbook_id",
    "category",
    "sheet_index",
    "sheet_id",
    "observed_unit",
    "canonical_unit",
    "dimension",
    "canonical_si_unit",
    "conversion_kind",
    "si_multiplier",
    "si_offset",
    "pressure_basis",
    "temperature_kind",
    "reference_state",
    "quantity_kind",
    "evidence_source",
    "evidence_confidence",
    "occurrences",
    "normalization_hazard",
]
DEFINED_NAME_INVENTORY_FIELDS = [
    "workbook_id",
    "category",
    "defined_name_id",
    "scope_sheet_index",
    "hidden",
    "metadata_status",
    "reference_kind",
    "external_reference",
    "functions",
    "calculation_topics",
    "detected_units",
]
VBA_INVENTORY_FIELDS = [
    "workbook_id",
    "category",
    "module_id",
    "record_id",
    "record_kind",
    "module_kind",
    "procedure_kind",
    "visibility",
    "execution_trigger",
    "source_status",
    "source_lines",
    "executable_lines",
    "calculation_topics",
    "detected_units",
    "risk_signals",
]
MACRO_INDICATOR_FIELDS = [
    "workbook_id",
    "category",
    "indicator_type",
    "risk_class",
    "occurrences",
]
PUBLIC_INDEX_FIELDS = [
    "workbook_id",
    "classification",
    "category",
    "extension",
    "bytes",
    "source_occurrences",
    "sheet_count",
    "formula_cells",
    "has_vba",
    "audit_stage",
    "audit_status",
    "static_coverage",
]


def _workbook_audit_row(
    document: dict[str, Any],
    sheet_rows: list[dict[str, Any]],
    calculation_rows: list[dict[str, Any]],
    unit_rows: list[dict[str, Any]],
    name_rows: list[dict[str, Any]],
    vba_rows: list[dict[str, Any]],
    indicator_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    metadata = document["workbook"]
    analysis = document["extraction"]
    macro_analysis = document["macro_analysis"]
    formula_cells = sum(int(row["formula_cells"]) for row in sheet_rows)
    topics = {
        topic
        for row in calculation_rows
        for topic in str(row["calculation_topics"]).split("|")
        if topic
    }
    topics.update(
        topic
        for row in vba_rows
        for topic in str(row["calculation_topics"]).split("|")
        if topic
    )
    extraction_status = str(analysis.get("status", ""))
    vba_status = str(macro_analysis.get("status", ""))
    method = str(analysis.get("method", ""))
    pcode_status = str(macro_analysis.get("pcode_inspection", "not-run"))
    if extraction_status == "ok" and vba_status == "ok":
        if pcode_status != "enabled":
            coverage = "formula-name-and-vba-source-only"
        elif method.startswith("calamine"):
            coverage = "calamine-supported-records-and-vba-complete"
        else:
            coverage = "ooxml-formulas-names-and-vba-complete"
    elif extraction_status == "ok" and vba_status == "skipped":
        coverage = "formula-only"
    elif extraction_status == "ok":
        coverage = "formula-complete-vba-partial"
    else:
        coverage = "partial"
    calculation = analysis.get("calculation", {})
    return {
        "workbook_id": str(metadata.get("id", "")),
        "category": str(metadata.get("category", "")),
        "extension": str(metadata.get("extension", "")),
        "bytes": int(metadata.get("bytes", 0) or 0),
        "integrity_matches_index": bool(
            document.get("integrity", {}).get("matches_index")
        ),
        "extraction_method": method,
        "extraction_status": extraction_status,
        "static_coverage": coverage,
        "sheet_count": len(sheet_rows),
        "hidden_sheet_count": sum(row["visibility"] == "hidden" for row in sheet_rows),
        "very_hidden_sheet_count": sum(
            row["visibility"] == "veryHidden" for row in sheet_rows
        ),
        "macro_sheet_count": sum(
            row["sheet_kind"] == "macro-sheet" for row in sheet_rows
        ),
        "formula_cells": formula_cells,
        "formula_families": len(calculation_rows),
        "defined_names": len(name_rows),
        "unit_dimensions": "|".join(sorted({row["dimension"] for row in unit_rows})),
        "canonical_units": "|".join(
            sorted({row["canonical_unit"] for row in unit_rows})
        ),
        "calculation_topics": "|".join(sorted(topics)),
        "external_reference_formulas": sum(
            int(row["external_reference_count"]) for row in calculation_rows
        ),
        "volatile_formula_cells": sum(
            int(row["occurrence_count"])
            for row in calculation_rows
            if row["volatile_functions"]
        ),
        "iterative_calculation": calculation.get("iterate", ""),
        "vba_status": vba_status,
        "vba_pcode_inspection": pcode_status,
        "has_vba": macro_analysis.get("has_vba", ""),
        "vba_modules": sum(row["record_kind"] == "module" for row in vba_rows),
        "vba_procedures": sum(row["record_kind"] == "procedure" for row in vba_rows),
        "event_procedures": sum(
            row["record_kind"] == "procedure"
            and row["execution_trigger"] != "explicit-call"
            for row in vba_rows
        ),
        "macro_indicator_count": sum(int(row["occurrences"]) for row in indicator_rows),
    }


def _open_csv_writer(path: Path, fieldnames: list[str]) -> tuple[Any, csv.DictWriter]:
    temporary = path.with_suffix(path.suffix + ".tmp")
    handle = temporary.open("w", newline="", encoding="utf-8")
    writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    return handle, writer


def _public_index_row(
    metadata: dict[str, Any], workbook_row: dict[str, Any]
) -> dict[str, Any]:
    return {
        "workbook_id": workbook_row["workbook_id"],
        "classification": str(metadata.get("classification", "drilling_calculation")),
        "category": workbook_row["category"],
        "extension": workbook_row["extension"],
        "bytes": workbook_row["bytes"],
        "source_occurrences": int(metadata.get("source_occurrences", 1) or 1),
        "sheet_count": workbook_row["sheet_count"],
        "formula_cells": workbook_row["formula_cells"],
        "has_vba": workbook_row["has_vba"],
        "audit_stage": "deep-static",
        "audit_status": "ok"
        if workbook_row["extraction_status"] == "ok"
        and workbook_row["vba_status"] == "ok"
        else "partial",
        "static_coverage": workbook_row["static_coverage"],
    }


def merge_captures(
    output_root: Path,
    public_dir: Path,
    expected_workbook_ids: set[str] | None = None,
    allow_partial: bool = False,
) -> dict[str, Any]:
    """Stream private per-workbook JSON into privacy-safe public inventories."""

    output_root = output_root.resolve()
    public_dir = public_dir.resolve()
    public_dir.mkdir(parents=True, exist_ok=True)
    capture_paths = sorted((output_root / "workbooks").glob("*.json"))
    if not capture_paths:
        raise ValueError("No private workbook captures are available to merge")
    capture_ids = {path.stem.casefold() for path in capture_paths}
    if any(not WORKBOOK_ID_PATTERN.fullmatch(item) for item in capture_ids):
        raise ValueError("Capture directory contains an invalid workbook id")
    if expected_workbook_ids is not None:
        expected = {item.casefold() for item in expected_workbook_ids}
        if capture_ids != expected:
            raise ValueError(
                "Capture/catalog reconciliation failed: "
                f"expected {len(expected)}, found {len(capture_ids)}"
            )
    definitions = {
        "INDEX.csv": PUBLIC_INDEX_FIELDS,
        "WORKBOOK_AUDIT.csv": WORKBOOK_AUDIT_FIELDS,
        "SHEET_INVENTORY.csv": SHEET_INVENTORY_FIELDS,
        "CALCULATION_INVENTORY.csv": CALCULATION_INVENTORY_FIELDS,
        "UNIT_INVENTORY.csv": UNIT_INVENTORY_FIELDS,
        "DEFINED_NAME_INVENTORY.csv": DEFINED_NAME_INVENTORY_FIELDS,
        "VBA_INVENTORY.csv": VBA_INVENTORY_FIELDS,
        "MACRO_INDICATORS.csv": MACRO_INDICATOR_FIELDS,
    }
    handles: dict[str, Any] = {}
    writers: dict[str, csv.DictWriter] = {}
    for filename, fieldnames in definitions.items():
        handle, writer = _open_csv_writer(public_dir / filename, fieldnames)
        handles[filename] = handle
        writers[filename] = writer

    summary: dict[str, Any] = {
        "workbooks": 0,
        "sheets": 0,
        "formula_cells": 0,
        "formula_families": 0,
        "unit_inventory_rows": 0,
        "defined_names": 0,
        "vba_modules": 0,
        "vba_procedures": 0,
        "macro_indicators": 0,
        "extraction_errors": 0,
        "vba_analysis_failures": 0,
        "integrity_failures": 0,
        "external_reference_formulas": 0,
        "volatile_formula_cells": 0,
        "unit_evidence_occurrences": 0,
        "workbooks_with_vba": 0,
        "workbooks_with_external_references": 0,
        "workbooks_with_volatile_formulas": 0,
        "formats": Counter(),
        "categories": Counter(),
        "extraction_methods": Counter(),
        "topics": Counter(),
        "units": Counter(),
        "vba_procedure_risk_records": Counter(),
    }
    completed = False
    try:
        for capture_path in capture_paths:
            document = _load_json_bounded(capture_path, MAX_READER_JSON_BYTES)
            if (
                not isinstance(document, dict)
                or document.get("analysis_schema_version") != "1.0.0"
            ):
                raise ValueError(f"Invalid capture schema for {capture_path.stem}")
            metadata = document["workbook"]
            analysis = document["extraction"]
            macro_analysis = document["macro_analysis"]
            if str(metadata.get("id", "")).casefold() != capture_path.stem.casefold():
                raise ValueError(f"Capture identity mismatch for {capture_path.stem}")
            integrity_ok = bool(document.get("integrity", {}).get("matches_index"))
            extraction_ok = analysis.get("status") == "ok"
            vba_status = str(macro_analysis.get("status", ""))
            vba_ok = vba_status in {"ok", "skipped"}
            pcode_ok = (
                vba_status == "skipped"
                or macro_analysis.get("pcode_inspection") == "enabled"
            )
            if not allow_partial and not (
                integrity_ok and extraction_ok and vba_ok and pcode_ok
            ):
                raise ValueError(f"Incomplete private capture for {capture_path.stem}")
            sheet_rows = build_sheet_rows(metadata, analysis)
            calculation_rows = build_calculation_rows(metadata, analysis)
            unit_rows = build_unit_rows(metadata, analysis)
            name_rows = build_defined_name_rows(metadata, analysis)
            vba_rows = build_vba_rows(metadata, macro_analysis)
            indicator_rows = build_macro_indicator_rows(metadata, macro_analysis)
            workbook_rows = [
                _workbook_audit_row(
                    document,
                    sheet_rows,
                    calculation_rows,
                    unit_rows,
                    name_rows,
                    vba_rows,
                    indicator_rows,
                )
            ]
            index_rows = [_public_index_row(metadata, workbook_rows[0])]
            row_sets = {
                "INDEX.csv": index_rows,
                "WORKBOOK_AUDIT.csv": workbook_rows,
                "SHEET_INVENTORY.csv": sheet_rows,
                "CALCULATION_INVENTORY.csv": calculation_rows,
                "UNIT_INVENTORY.csv": unit_rows,
                "DEFINED_NAME_INVENTORY.csv": name_rows,
                "VBA_INVENTORY.csv": vba_rows,
                "MACRO_INDICATORS.csv": indicator_rows,
            }
            for filename, rows in row_sets.items():
                assert_public_records_safe(rows)
                writers[filename].writerows(rows)

            workbook_row = workbook_rows[0]
            summary["workbooks"] += 1
            summary["sheets"] += len(sheet_rows)
            summary["formula_cells"] += int(workbook_row["formula_cells"])
            summary["formula_families"] += len(calculation_rows)
            summary["unit_inventory_rows"] += len(unit_rows)
            summary["defined_names"] += len(name_rows)
            summary["vba_modules"] += int(workbook_row["vba_modules"])
            summary["vba_procedures"] += int(workbook_row["vba_procedures"])
            summary["macro_indicators"] += int(workbook_row["macro_indicator_count"])
            summary["external_reference_formulas"] += int(
                workbook_row["external_reference_formulas"]
            )
            summary["volatile_formula_cells"] += int(
                workbook_row["volatile_formula_cells"]
            )
            summary["unit_evidence_occurrences"] += sum(
                int(row["occurrences"]) for row in unit_rows
            )
            summary["workbooks_with_vba"] += bool(workbook_row["has_vba"])
            summary["workbooks_with_external_references"] += (
                int(workbook_row["external_reference_formulas"]) > 0
            )
            summary["workbooks_with_volatile_formulas"] += (
                int(workbook_row["volatile_formula_cells"]) > 0
            )
            summary["extraction_errors"] += analysis.get("status") != "ok"
            summary["vba_analysis_failures"] += vba_status not in {"ok", "skipped"}
            summary["integrity_failures"] += not document.get("integrity", {}).get(
                "matches_index", False
            )
            summary["formats"][str(metadata.get("extension", ""))] += 1
            summary["categories"][str(metadata.get("category", ""))] += 1
            summary["extraction_methods"][str(analysis.get("method", ""))] += 1
            for row in calculation_rows:
                for topic in str(row["calculation_topics"]).split("|"):
                    if topic:
                        summary["topics"][topic] += int(row["occurrence_count"])
            for row in unit_rows:
                summary["units"][row["canonical_unit"]] += int(row["occurrences"])
            for row in vba_rows:
                if row["record_kind"] != "procedure":
                    continue
                for signal in str(row["risk_signals"]).split("|"):
                    if signal:
                        summary["vba_procedure_risk_records"][signal] += 1

            del (
                document,
                metadata,
                analysis,
                macro_analysis,
                sheet_rows,
                calculation_rows,
                unit_rows,
                name_rows,
                vba_rows,
                indicator_rows,
                workbook_rows,
                index_rows,
                row_sets,
            )
            import gc

            gc.collect()
        completed = True
    finally:
        for handle in handles.values():
            handle.close()
        if not completed:
            for filename in definitions:
                (public_dir / filename).with_suffix(
                    (public_dir / filename).suffix + ".tmp"
                ).unlink(missing_ok=True)

    if completed:
        for filename in definitions:
            destination = public_dir / filename
            os.replace(
                destination.with_suffix(destination.suffix + ".tmp"), destination
            )

    serializable_summary = {
        key: dict(sorted(value.items())) if isinstance(value, Counter) else value
        for key, value in summary.items()
    }
    _write_json_atomic(output_root / "MERGE_SUMMARY.json", serializable_summary)
    _write_json_atomic(public_dir / "ANALYSIS_SUMMARY.json", serializable_summary)
    return serializable_summary


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--catalog",
        type=Path,
        default=repo_root / "research" / "drilling-calculation-workbooks",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=repo_root / "outputs" / "drilling-workbook-analysis",
        help="Ignored directory containing one private JSON file per workbook.",
    )
    parser.add_argument(
        "--public-dir",
        type=Path,
        default=repo_root / "research" / "drilling-calculation-workbooks",
        help="Directory receiving privacy-filtered merged CSV inventories.",
    )
    parser.add_argument(
        "--static-reader",
        type=Path,
        help="Path to the wellforge-workbook-audit executable.",
    )
    parser.add_argument(
        "--olevba",
        type=Path,
        help="Path to olevba.exe; auto-detected when omitted.",
    )
    parser.add_argument(
        "--msoffcrypto",
        type=Path,
        help="Path to msoffcrypto-tool; auto-detected when omitted.",
    )
    parser.add_argument("--timeout", type=int, default=300)
    parser.add_argument("--skip-vba", action="store_true")
    parser.add_argument(
        "--allow-partial",
        action="store_true",
        help="Publish incomplete captures instead of failing closed.",
    )
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--limit", type=int)
    parser.add_argument(
        "--workbook-id",
        action="append",
        dest="workbook_ids",
        help="Capture only this catalog id; may be supplied more than once.",
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--capture-only", action="store_true")
    mode.add_argument("--merge-only", action="store_true")
    parser.add_argument("--internal-openxml-input", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--internal-openxml-output", type=Path, help=argparse.SUPPRESS)
    return parser.parse_args()


def _default_static_reader(repo_root: Path) -> Path | None:
    executable = (
        "wellforge-workbook-audit.exe"
        if os.name == "nt"
        else "wellforge-workbook-audit"
    )
    for profile in ("release", "debug"):
        candidate = repo_root / "engine" / "target" / profile / executable
        if candidate.is_file():
            return candidate
    return None


def _default_olevba() -> Path | None:
    discovered = shutil.which("olevba")
    if discovered:
        return Path(discovered)
    known = Path(r"C:\Program Files\Python313\Scripts\olevba.exe")
    return known if known.is_file() else None


def _default_msoffcrypto() -> Path | None:
    executable = "msoffcrypto-tool.exe" if os.name == "nt" else "msoffcrypto-tool"
    discovered = shutil.which(executable)
    if discovered:
        return Path(discovered)
    known = Path(r"C:\Program Files\Python313\Scripts\msoffcrypto-tool.exe")
    return known if known.is_file() else None


def main() -> int:
    args = parse_args()
    if (
        args.internal_openxml_input is not None
        or args.internal_openxml_output is not None
    ):
        if args.internal_openxml_input is None or args.internal_openxml_output is None:
            raise ValueError("Both internal OOXML paths are required")
        extraction = compact_formula_families(
            extract_ooxml_workbook(args.internal_openxml_input.resolve(strict=True))
        )
        _write_json_atomic(args.internal_openxml_output.resolve(), extraction)
        return 0
    repo_root = Path(__file__).resolve().parents[1]
    static_reader = args.static_reader or _default_static_reader(repo_root)
    olevba_path = args.olevba or _default_olevba()
    msoffcrypto_path = args.msoffcrypto or _default_msoffcrypto()
    result: dict[str, Any] = {}
    if not args.merge_only:
        result["capture"] = capture_catalog(
            catalog=args.catalog,
            output_root=args.output_root,
            static_reader=static_reader,
            olevba_path=olevba_path,
            skip_vba=args.skip_vba,
            timeout_seconds=args.timeout,
            force=args.force,
            limit=args.limit,
            workbook_ids=set(args.workbook_ids or []),
            msoffcrypto_path=msoffcrypto_path,
        )
        if result["capture"]["errors"] and not args.allow_partial:
            raise RuntimeError(
                f"{result['capture']['errors']} workbook captures are incomplete"
            )
    if not args.capture_only:
        catalog_records = _load_catalog_records(args.catalog.resolve())
        result["merge"] = merge_captures(
            output_root=args.output_root,
            public_dir=args.public_dir,
            expected_workbook_ids={record["id"] for record in catalog_records},
            allow_partial=args.allow_partial,
        )
    print(json.dumps(result, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
